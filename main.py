"""
MF Holdings API v7 — Universal AMC Parser + ISIN-based cap classification
Formats confirmed:
  A: Sundaram/Nippon/Axis/ABSL — Index + fund sheets, % decimal
  B: SBI — Index + fund sheets, % actual
  C: Kotak — No Index, name offset in data rows, % actual
  D: ICICI/HDFC — One xlsx per fund (ZIP of individual files)
  E: UTI — Single sheet, SCHEME CODE###STARTS/ENDS markers
"""

import os, re, io, logging, json, zipfile, httpx, asyncio, time
from datetime import datetime, date
from typing import Optional, List
from pathlib import Path

import openpyxl
from fastapi import FastAPI, Query, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app):
    load_db()
    yield

app = FastAPI(title="MF Holdings API", version="7.0.0", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["GET","POST","DELETE"], allow_headers=["*"])

# ════════════════════════════════════════════════════════════════
# PERSISTENCE -- Firestore-backed, not local filesystem.
#
# Render's filesystem (including /tmp) is EPHEMERAL -- it is wiped on
# every service restart, which happens routinely (free tier sleep/wake,
# every redeploy, periodic maintenance even on paid tiers without an
# explicit persistent disk attached). Storing holdings_db as a local
# JSON file meant every upload appeared to succeed in the moment, but
# was silently lost on the next restart -- the server would keep
# serving whatever snapshot happened to survive, while new uploads
# never durably accumulated. This was a real, confirmed bug (server
# stuck at 768 funds / a stale April timestamp despite many uploads).
#
# Fix: persist holdings_db to Firestore instead, sharded across
# multiple documents (Firestore's 1MB-per-document limit means the
# full 768+ fund dataset cannot fit in a single doc -- same constraint
# already solved on the client side for AMC factsheet data earlier in
# this project). A local JSON cache is still kept as a fast in-memory
# mirror and a fallback if Firestore is temporarily unreachable, but
# Firestore is now the source of truth.
# ════════════════════════════════════════════════════════════════
DATA_DIR = Path(os.environ.get("DATA_DIR", "/tmp/mf_data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_FILE  = DATA_DIR / "holdings.json"  # local fallback cache only, not source of truth
holdings_db: dict = {}
_amfi_cap_cache: dict = {}
_firestore_db = None  # lazily initialized Firestore client, None if unavailable

FIRESTORE_SHARD_SIZE = 40  # funds per shard document, tuned to stay well under 1MB

def _get_firestore_client():
    """Lazily initialize the Firebase Admin SDK / Firestore client from a
    service account JSON provided via the FIREBASE_SERVICE_ACCOUNT_JSON
    environment variable (set in Render's dashboard, never committed to
    the repo). Returns None if not configured or initialization fails --
    callers must handle that gracefully and fall back to local cache.
    """
    global _firestore_db
    if _firestore_db is not None:
        return _firestore_db
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore as fb_firestore
        sa_json = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", "")
        if not sa_json:
            log.warning("FIREBASE_SERVICE_ACCOUNT_JSON not set -- holdings_db will NOT "
                        "persist across server restarts. Set this env var in Render to fix.")
            return None
        if not firebase_admin._apps:
            cred = credentials.Certificate(json.loads(sa_json))
            firebase_admin.initialize_app(cred)
        _firestore_db = fb_firestore.client()
        log.info("Firestore client initialized successfully")
        return _firestore_db
    except Exception as e:
        log.error(f"Firestore initialization failed: {e}. holdings_db will NOT persist.")
        return None

# ---------------------------------------------------------------------------
# AMFI CAP MAP — loaded once from bundled Excel at startup
# Keys: ISIN strings. Values: 'large' | 'mid'  (absent = 'small')
# File: amfi_market_cap.xlsx in the same directory as main.py
# Update every Jan / Jul: replace the file and redeploy.
# ---------------------------------------------------------------------------
def _load_amfi_cap_map_from_file(path: str = "amfi_market_cap.xlsx") -> dict:
    if not os.path.exists(path):
        log.warning(f"AMFI cap file not found at '{path}'. Will fall back to live fetch.")
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        cap_map = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            isin     = row[2]   # col C
            category = row[10]  # col K
            if not isin or not category:
                continue
            cat = str(category).strip().lower()
            if "large" in cat:
                cap_map[isin.strip()] = "large"
            elif "mid" in cat:
                cap_map[isin.strip()] = "mid"
            # small cap entries skipped — default is 'small'
        wb.close()
        large_n = sum(1 for v in cap_map.values() if v == "large")
        mid_n   = sum(1 for v in cap_map.values() if v == "mid")
        log.info(f"AMFI cap map loaded from file: {large_n} large, {mid_n} mid ({len(cap_map)} total)")
        return cap_map
    except Exception as e:
        log.warning(f"AMFI cap file load failed: {e}. Will fall back to live fetch.")
        return {}

AMFI_ISIN_CAP: dict = _load_amfi_cap_map_from_file()

# ---------------------------------------------------------------------------
# ISIN OVERRIDE MAP — stocks where AMFI Excel has old ISIN but AMC disclosures
# use a new ISIN (issued after bonus shares / stock splits / reclassification).
# Add new entries here as discovered. Format: { new_isin: 'large'|'mid'|'small' }
# ---------------------------------------------------------------------------
ISIN_OVERRIDES: dict = {
    "INE1TAE01010": "large",   # Tata Motors (new ISIN, old: INE155A01022)
    "INE237A01036": "large",   # Kotak Mahindra Bank (new ISIN, old: INE237A01028)
    "INE745G01043": "mid",     # Multi Commodity Exchange (new ISIN, old: INE745G01035)
}
AMFI_ISIN_CAP.update(ISIN_OVERRIDES)
log.info(f"ISIN overrides applied: {len(ISIN_OVERRIDES)} entries")

def save_db():
    """Persist holdings_db to Firestore (sharded across multiple documents
    to stay under the 1MB-per-document limit), plus a local JSON cache as
    a fast fallback. Firestore is the durable store; the local file is
    disposable and only used to avoid a network round-trip on every read
    within a single server lifetime.
    """
    try:
        DB_FILE.write_text(json.dumps(holdings_db, ensure_ascii=False))
    except Exception as e:
        log.warning(f"Local cache save failed (non-fatal): {e}")

    db = _get_firestore_client()
    if db is None:
        log.warning("Firestore unavailable -- holdings_db saved locally only, "
                     "WILL BE LOST on next restart")
        return

    try:
        keys = sorted(holdings_db.keys())
        shards = [keys[i:i + FIRESTORE_SHARD_SIZE] for i in range(0, len(keys), FIRESTORE_SHARD_SIZE)]
        batch = db.batch()
        batch_count = 0
        for shard_idx, shard_keys in enumerate(shards):
            shard_data = {k: holdings_db[k] for k in shard_keys}
            doc_ref = db.collection("mf_holdings_shards").document(f"shard_{shard_idx}")
            batch.set(doc_ref, {"keys": shard_keys, "data": shard_data})
            batch_count += 1
            # Firestore batches cap at 500 operations -- commit and start a
            # fresh batch well before that if a future dataset grows huge
            if batch_count >= 400:
                batch.commit()
                batch = db.batch()
                batch_count = 0
        if batch_count > 0:
            batch.commit()
        # Clean up any stale shard documents from a previous save that had
        # MORE shards than this one (e.g. after bulk deletes) -- otherwise
        # old shard docs linger forever with stale data that load_db()
        # would incorrectly merge back in on next startup
        meta_ref = db.collection("mf_holdings_shards").document("_meta")
        prev_meta = meta_ref.get()
        prev_shard_count = prev_meta.to_dict().get("shard_count", 0) if prev_meta.exists else 0
        if prev_shard_count > len(shards):
            cleanup_batch = db.batch()
            for i in range(len(shards), prev_shard_count):
                cleanup_batch.delete(db.collection("mf_holdings_shards").document(f"shard_{i}"))
            cleanup_batch.commit()
        meta_ref.set({"shard_count": len(shards), "total_funds": len(holdings_db),
                      "updated_at": datetime.utcnow().isoformat()})
        log.info(f"Saved {len(holdings_db)} funds to Firestore across {len(shards)} shards")
    except Exception as e:
        log.error(f"Firestore save FAILED: {e} -- holdings_db only saved locally, "
                  f"WILL BE LOST on next restart")

def load_db():
    """Load holdings_db from Firestore (source of truth). Falls back to
    the local JSON cache only if Firestore is unreachable -- which means
    falling back to whatever happened to survive on local disk, the same
    fragile behavior we're moving away from, but better than starting
    completely empty.
    """
    global holdings_db
    db = _get_firestore_client()
    if db is not None:
        try:
            shard_docs = db.collection("mf_holdings_shards").stream()
            loaded = {}
            shard_count = 0
            for doc in shard_docs:
                if doc.id == "_meta":
                    continue
                shard_data = doc.to_dict().get("data", {})
                loaded.update(shard_data)
                shard_count += 1
            if loaded:
                holdings_db = loaded
                log.info(f"Loaded {len(holdings_db)} funds from Firestore ({shard_count} shards)")
                try:
                    DB_FILE.write_text(json.dumps(holdings_db, ensure_ascii=False))
                except Exception:
                    pass
                return
            log.info("Firestore reachable but no holdings data found yet (fresh database)")
            return
        except Exception as e:
            log.error(f"Firestore load FAILED: {e} -- falling back to local cache")

    # Firestore unavailable or empty -- fall back to local file (fragile,
    # may be stale or missing entirely after a restart)
    try:
        if DB_FILE.exists():
            holdings_db = json.loads(DB_FILE.read_text())
            log.warning(f"Loaded {len(holdings_db)} funds from LOCAL CACHE ONLY -- "
                        f"Firestore was unavailable, this data may be stale")
    except Exception as e:
        log.warning(f"Local cache load failed: {e}")

# AMC UPLOAD AGENT -- Plan, Act, Verify, Repair, Report
# Same shape as the market-data agent, applied to stock classification
# gaps in uploaded AMC disclosures:
#   1. PLAN   -- every equity holding should resolve a real cap
#                classification (large/mid/small) via ISIN or name match
#   2. ACT    -- the existing parser + _enrich_holdings already does this
#   3. VERIFY -- find holdings that fell through to the silent "small"
#                default because neither ISIN nor name matched anything
#   4. REPAIR -- one cached AI lookup per genuinely-unresolved stock,
#                resolved once, cached permanently by ISIN (or normalized
#                name if ISIN is missing), so the same stock is never
#                billed for twice across uploads
#   5. REPORT -- a per-upload health record, kept in a rolling log and
#                viewable via /amc-health

CAP_CACHE_FILE = DATA_DIR / "cap_resolution_cache.json"
_cap_resolution_cache: dict = {}

def _load_cap_cache():
    global _cap_resolution_cache
    try:
        if CAP_CACHE_FILE.exists():
            _cap_resolution_cache = json.loads(CAP_CACHE_FILE.read_text())
            log.info(f"Cap resolution cache loaded: {len(_cap_resolution_cache)} entries")
    except Exception as e:
        log.warning(f"Cap cache load failed: {e}")

def _save_cap_cache():
    try:
        CAP_CACHE_FILE.write_text(json.dumps(_cap_resolution_cache, ensure_ascii=False))
    except Exception as e:
        log.warning(f"Cap cache save failed: {e}")

_load_cap_cache()

_amc_health: list = []
_AMC_HEALTH_MAX = 30

def _amc_health_log(report: dict):
    report["ts"] = time.time()
    report["time_str"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _amc_health.insert(0, report)
    while len(_amc_health) > _AMC_HEALTH_MAX:
        _amc_health.pop()

def _agent_verify_amc(fund_data: dict, enriched_holdings: list) -> dict:
    """VERIFY step for one fund -- find equity holdings whose cap classification
    fell through to the silent 'small' default with no real match behind it
    (no ISIN hit, no AMFI name hit, no override, no cache hit)."""
    unresolved = []
    cap_map = _amfi_cap_map()
    for h in enriched_holdings:
        if h.get("type") == "debt":
            continue
        isin = (h.get("isin") or "").strip()
        name = h.get("name") or ""
        real_match = False
        if isin and AMFI_ISIN_CAP.get(isin) is not None:
            real_match = True
        elif isin and _cap_resolution_cache.get(isin, {}).get("cap"):
            real_match = True
        else:
            key = _norm_stock(name)
            if cap_map.get(key) is not None:
                real_match = True
            elif _cap_resolution_cache.get(key, {}).get("cap"):
                real_match = True
        if not real_match:
            unresolved.append({"name": name, "isin": isin})
    return {"ok": not unresolved, "unresolved": unresolved}

async def _agent_repair_cap(api_key: str, stock_name: str, isin: str = "") -> Optional[str]:
    """REPAIR step -- resolve ONE stock's market-cap category via Claude + web
    search, then cache the result permanently keyed by ISIN (preferred) or
    normalized name. Returns 'large' | 'mid' | 'small' | None."""
    cache_key = isin if isin else _norm_stock(stock_name)
    cached = _cap_resolution_cache.get(cache_key)
    if cached and cached.get("cap"):
        return cached["cap"]

    if not api_key:
        return None

    prompt = (
        f'What is the SEBI/AMFI market capitalisation category of the Indian listed '
        f'company "{stock_name}"{f" (ISIN {isin})" if isin else ""} -- is it Large Cap '
        f'(top 100 by market cap), Mid Cap (101-250), or Small Cap (251+)? '
        f'Search the web for its current market capitalisation and AMFI classification. '
        f'Return ONLY one word: "large", "mid", or "small". If you cannot determine it '
        f'with reasonable confidence after searching, return "unknown".'
    )
    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 200,
                    "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 2}],
                    "messages": [{"role": "user", "content": prompt}]
                })
            if resp.status_code != 200:
                log.warning(f"Cap repair HTTP {resp.status_code} for '{stock_name}'")
                return None
            data = resp.json()
            blocks = data.get("content", [])
            if data.get("stop_reason") == "tool_use":
                tool_results = [
                    {"type": "tool_result", "tool_use_id": b.get("tool_use_id"), "content": b.get("content", [])}
                    for b in blocks if b.get("type") == "web_search_tool_result"
                ]
                resp2 = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 100,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 2}],
                        "messages": [
                            {"role": "user", "content": prompt},
                            {"role": "assistant", "content": blocks},
                            {"role": "user", "content": tool_results if tool_results else
                             [{"type": "text", "text": "Answer now with one word based on the search results."}]}
                        ]
                    })
                blocks = resp2.json().get("content", []) if resp2.status_code == 200 else blocks
            text = " ".join(b.get("text", "") for b in blocks if b.get("type") == "text").strip().lower()
            cap = next((c for c in ("large", "mid", "small") if c in text), None)
            if cap:
                _cap_resolution_cache[cache_key] = {
                    "cap": cap, "name": stock_name, "source": "ai_repair", "resolved_at": time.time()
                }
                _save_cap_cache()
                log.info(f"Agent repair resolved '{stock_name}' -> {cap}")
            return cap
    except Exception as e:
        log.warning(f"Cap repair failed for '{stock_name}': {e}")
    return None

def _amc_rename_candidates(amc_name: str, new_key: str, new_fund_name: str) -> list:
    """Find existing holdings_db entries for the SAME AMC whose normalized
    key is suspiciously similar to a brand-new key, but not identical --
    the signature of an AMC renaming a scheme between monthly disclosures
    (e.g. dropping/adding a word, changing 'Fund' to 'Scheme', etc).
    Pure local heuristic, no AI cost for this screening step."""
    candidates = []
    new_words = set(w for w in new_key.split() if len(w) >= 3)
    if not new_words:
        return candidates
    for existing_key, existing_data in holdings_db.items():
        if existing_key == new_key:
            continue  # exact match -- not a rename, handled by normal overwrite
        if existing_data.get("amc", "").strip().lower() != amc_name.strip().lower():
            continue  # only consider renames within the same AMC
        existing_words = set(w for w in existing_key.split() if len(w) >= 3)
        if not existing_words:
            continue
        overlap = len(new_words & existing_words)
        union_size = len(new_words | existing_words)
        similarity = overlap / union_size if union_size else 0
        # High overlap (most words shared) but not identical -- rename signature
        if similarity >= 0.55:
            candidates.append({
                "key": existing_key,
                "fund_name": existing_data.get("fund_name", existing_key),
                "similarity": round(similarity, 2),
            })
    candidates.sort(key=lambda c: -c["similarity"])
    return candidates[:3]  # at most 3 candidates worth asking AI about

async def _agent_confirm_rename(api_key: str, amc_name: str, new_name: str, old_name: str) -> bool:
    """AI judgment call -- is `new_name` actually the same scheme as
    `old_name`, just renamed by the AMC, or are these genuinely two
    different funds that happen to share similar words? Cached by the
    (old,new) pair so the same rename decision is never billed twice."""
    cache_key = f"{norm(old_name)}|||{norm(new_name)}"
    if cache_key in _cas_resolve_cache:
        cached = _cas_resolve_cache[cache_key]
        return cached.get("is_rename", False)

    if not api_key:
        return False  # no AI available -- do not guess, leave both entries as-is

    prompt = (
        f'AMC "{amc_name}" has an existing fund in our database called '
        f'"{old_name}". A new monthly disclosure upload contains a fund called '
        f'"{new_name}" that was not in our database before.\n\n'
        f'Is "{new_name}" actually the SAME scheme as "{old_name}", just renamed '
        f'by the AMC (e.g. SEBI-mandated rename, AMC merger, scheme repositioning)? '
        f'Or are these genuinely two different, separate funds?\n\n'
        f'If you are not highly confident it is a rename of the same scheme, answer NO -- '
        f'it is safer to treat them as different funds than to wrongly merge two distinct schemes.\n\n'
        f'Return ONLY one word: "YES" or "NO".'
    )
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 10,
                    "messages": [{"role": "user", "content": prompt}]
                })
            if resp.status_code != 200:
                return False
            text = resp.json().get("content", [{}])[0].get("text", "").strip().upper()
            is_rename = text.startswith("YES")
            _cas_resolve_cache[cache_key] = {"is_rename": is_rename, "checked_at": time.time()}
            _save_cas_resolve_cache()
            return is_rename
    except Exception as e:
        log.warning(f"Rename confirmation failed for '{new_name}' vs '{old_name}': {e}")
        return False

async def _detect_and_resolve_renames(amc_name: str, parsed: dict, api_key: str) -> dict:
    """Runs BEFORE holdings_db.update() -- for every brand-new fund key in
    this upload, check if it's likely a rename of an existing fund for the
    same AMC. If AI confirms with high confidence, delete the OLD key so
    the new one cleanly replaces it instead of sitting alongside as a
    silent duplicate. Fully automatic -- no manual decision required.
    Returns a log of what was auto-merged, for visibility in /amc-health.
    """
    renames_resolved = []
    for new_key, fund_data in parsed.items():
        if new_key in holdings_db:
            continue  # exact key match -- normal overwrite, not a rename case
        candidates = _amc_rename_candidates(amc_name, new_key, fund_data.get("fund_name", new_key))
        for cand in candidates:
            confirmed = await _agent_confirm_rename(
                api_key, amc_name, fund_data.get("fund_name", new_key), cand["fund_name"]
            )
            if confirmed:
                del holdings_db[cand["key"]]
                renames_resolved.append({
                    "old_name": cand["fund_name"], "old_key": cand["key"],
                    "new_name": fund_data.get("fund_name", new_key), "new_key": new_key,
                    "similarity": cand["similarity"],
                })
                log.info(f"Rename auto-resolved: '{cand['fund_name']}' -> "
                         f"'{fund_data.get('fund_name', new_key)}'")
                break  # only merge into the single best candidate
    return {"renames_resolved": renames_resolved}

async def _run_amc_upload_agent(amc_name: str, parsed_funds: dict, api_key: str = "", rename_report: dict = None):
    """Full agent cycle for one AMC upload -- verify every fund's enriched
    holdings, repair genuinely unresolved stocks (capped per upload to keep
    cost bounded), and log a per-AMC health report."""
    report = {"amc": amc_name, "funds": [], "total_unresolved": 0, "total_repaired": 0,
               "renames_resolved": (rename_report or {}).get("renames_resolved", [])}
    MAX_REPAIRS_PER_UPLOAD = 15
    repairs_done = 0

    for key, fund_data in parsed_funds.items():
        enriched = _enrich_holdings(fund_data)["holdings"]
        verdict = _agent_verify_amc(fund_data, enriched)
        fund_report = {
            "fund_name": fund_data.get("fund_name", key),
            "holdings_count": len(enriched),
            "unresolved_count": len(verdict["unresolved"]),
            "repaired": [],
        }
        report["total_unresolved"] += len(verdict["unresolved"])

        if verdict["unresolved"] and api_key and repairs_done < MAX_REPAIRS_PER_UPLOAD:
            for item in verdict["unresolved"]:
                if repairs_done >= MAX_REPAIRS_PER_UPLOAD:
                    break
                cap = await _agent_repair_cap(api_key, item["name"], item["isin"])
                repairs_done += 1
                if cap and cap != "unknown":
                    fund_report["repaired"].append({"name": item["name"], "cap": cap})
                    report["total_repaired"] += 1

        report["funds"].append(fund_report)

    report["repairs_used"] = repairs_done
    report["repairs_capped"] = repairs_done >= MAX_REPAIRS_PER_UPLOAD
    _amc_health_log(report)
    log.info(f"AMC agent [{amc_name}]: {report['total_unresolved']} unresolved, "
             f"{report['total_repaired']} repaired via AI, {repairs_done} repair calls used")
    return report

# ════════════════════════════════════════════════════════════════
# CAS IMPORT AGENT -- Pass 3: AI-assisted fund resolution
#
# Pass 1 (client-side exact/significant-word match against local
# factsheets) and Pass 2 (server /search with a 0.4 score threshold)
# already run client-side before this is ever called. This endpoint
# is Pass 3 -- the AI repair step for funds that survived both and
# are still unmatched, due to genuine name variation (AMC renamed a
# fund, merged schemes, abbreviation the local rules don't cover,
# OCR noise from a scanned CAS PDF, etc).
#
# Resolution is cached permanently by the exact CAS-stated fund name
# string, so the same fund name is never billed for AI resolution
# twice across any client's CAS import.
# ════════════════════════════════════════════════════════════════

CAS_RESOLVE_CACHE_FILE = DATA_DIR / "cas_resolve_cache.json"
_cas_resolve_cache: dict = {}

def _load_cas_resolve_cache():
    global _cas_resolve_cache
    try:
        if CAS_RESOLVE_CACHE_FILE.exists():
            _cas_resolve_cache = json.loads(CAS_RESOLVE_CACHE_FILE.read_text())
            log.info(f"CAS resolve cache loaded: {len(_cas_resolve_cache)} entries")
    except Exception as e:
        log.warning(f"CAS resolve cache load failed: {e}")

def _save_cas_resolve_cache():
    try:
        CAS_RESOLVE_CACHE_FILE.write_text(json.dumps(_cas_resolve_cache, ensure_ascii=False))
    except Exception as e:
        log.warning(f"CAS resolve cache save failed: {e}")

_load_cas_resolve_cache()

async def _agent_resolve_fund_name(api_key: str, cas_name: str, candidates: list) -> dict:
    """Ask Claude to decide whether one of the candidate funds (from /search)
    is actually the fund named in the CAS statement, accounting for AMC
    renames, abbreviations, OCR noise, and merged schemes. If none of the
    candidates are a real match, ask it to identify the fund directly via
    web search and return its current correct name for a follow-up search.

    Returns: {"action": "matched", "key": "...", "name": "...", "confidence": "high|medium"}
          or {"action": "rename_suggestion", "suggested_name": "...", "confidence": "..."}
          or {"action": "unresolved", "reason": "..."}
    """
    cache_key = norm(cas_name)
    if cache_key in _cas_resolve_cache:
        return _cas_resolve_cache[cache_key]

    if not api_key:
        return {"action": "unresolved", "reason": "no_api_key"}

    candidates_text = "\n".join(
        f'{i+1}. "{c["name"]}" (AMC: {c["amc"]}, search score: {c["score"]})'
        for i, c in enumerate(candidates[:10])
    ) or "(no candidates found in local database)"

    prompt = (
        f'A client\'s CAS (Consolidated Account Statement) lists a mutual fund holding as:\n'
        f'"{cas_name}"\n\n'
        f'Our database has these candidate funds that scored too low to auto-match '
        f'(threshold not met):\n{candidates_text}\n\n'
        f'Indian mutual fund names vary due to AMC mergers, scheme renames, plan/option '
        f'abbreviations, and CAS formatting differences. Decide:\n'
        f'1. Is the CAS name actually one of the candidates above, just named differently? '
        f'If yes, which one (give its exact number)?\n'
        f'2. If NONE of the candidates are a real match, search the web to identify what '
        f'fund "{cas_name}" actually refers to (check for AMC renames/mergers -- e.g. a '
        f'fund house may have changed its name or merged with another).\n\n'
        f'Return ONLY JSON, no other text:\n'
        f'{{"action":"matched","candidate_number":1,"confidence":"high"}} OR\n'
        f'{{"action":"rename_suggestion","suggested_name":"actual current fund name to search for","confidence":"medium"}} OR\n'
        f'{{"action":"unresolved","reason":"brief reason"}}'
    )

    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 300,
                    "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 2}],
                    "messages": [{"role": "user", "content": prompt}]
                })
            if resp.status_code != 200:
                log.warning(f"CAS resolve HTTP {resp.status_code} for '{cas_name}'")
                return {"action": "unresolved", "reason": f"api_error_{resp.status_code}"}
            data = resp.json()
            blocks = data.get("content", [])
            if data.get("stop_reason") == "tool_use":
                tool_results = [
                    {"type": "tool_result", "tool_use_id": b.get("tool_use_id"), "content": b.get("content", [])}
                    for b in blocks if b.get("type") == "web_search_tool_result"
                ]
                resp2 = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 250,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 2}],
                        "messages": [
                            {"role": "user", "content": prompt},
                            {"role": "assistant", "content": blocks},
                            {"role": "user", "content": tool_results if tool_results else
                             [{"type": "text", "text": "Provide the JSON decision now."}]}
                        ]
                    })
                blocks = resp2.json().get("content", []) if resp2.status_code == 200 else blocks
            text = " ".join(b.get("text", "") for b in blocks if b.get("type") == "text").strip()
            text = re.sub(r"```[^\n]*\n?|```", "", text).strip()
            start, end = text.find("{"), text.rfind("}") + 1
            if start < 0 or end <= start:
                return {"action": "unresolved", "reason": "no_json_in_response"}
            decision = json.loads(text[start:end])

            result = {"action": "unresolved", "reason": "unknown"}
            if decision.get("action") == "matched":
                idx = decision.get("candidate_number", 0) - 1
                if 0 <= idx < len(candidates[:10]):
                    c = candidates[idx]
                    result = {"action": "matched", "key": c["key"], "name": c["name"],
                              "amc": c["amc"], "confidence": decision.get("confidence", "medium")}
            elif decision.get("action") == "rename_suggestion":
                result = {"action": "rename_suggestion",
                          "suggested_name": decision.get("suggested_name", ""),
                          "confidence": decision.get("confidence", "low")}
            else:
                result = {"action": "unresolved", "reason": decision.get("reason", "ai_could_not_determine")}

            _cas_resolve_cache[cache_key] = result
            _save_cas_resolve_cache()
            log.info(f"CAS agent resolved '{cas_name}' -> {result['action']}")
            return result
    except Exception as e:
        log.warning(f"CAS resolve failed for '{cas_name}': {e}")
        return {"action": "unresolved", "reason": str(e)[:100]}

# ════════════════════════════════════════════════════════════════
# CAS EXTRACTION AGENT -- Plan, Act, Verify, Repair, Report
#
# Mirrors the AMC Upload Agent's shape exactly, applied to client
# portfolio statements (CAS/CAMS/KFintech/Scripbox/Groww/Kuvera/etc)
# instead of AMC factsheets:
#   1. PLAN   -- a correctly parsed statement should yield at least one
#                fund, with market values that reconcile against any
#                stated "Total"/"Grand Total" figure in the document
#   2. ACT    -- the deterministic client-side parser (CAMS-style regex,
#                Scripbox heuristic, etc) runs FIRST -- free, instant,
#                zero AI cost, same as Excel-first for AMC uploads
#   3. VERIFY -- the client sends its raw extracted text + what it
#                parsed (possibly nothing) to this endpoint; we check:
#                did it find >=1 fund? Do the values reconcile to within
#                5% of a detected total figure?
#   4. REPAIR -- if verification fails (new/unrecognized provider
#                layout, or totals don't reconcile), ask Claude to read
#                the raw statement text directly and extract every fund
#                holding, regardless of the provider's specific format.
#                This REPLACES hand-coding a new regex per provider --
#                any new statement layout (Groww, Kuvera, ETMoney,
#                INDmoney, a future CAMS template change) is handled by
#                the same repair path without needing new code.
#   5. REPORT -- which path was used (deterministic vs AI-repaired),
#                surfaced in the import preview so the advisor always
#                knows when AI extraction was needed.
# ════════════════════════════════════════════════════════════════

def _cas_verify_extraction(funds: list, raw_text: str) -> dict:
    """VERIFY step -- did the extraction actually work?

    Two independent checks:
    1. Reconciliation against the document's own "Grand Total" row, being
       careful to pick the MKT VALUE figure specifically -- the Grand
       Total row has multiple numbers (Cost, Mkt Value, Unrealized G/L,
       Total Returns...) in sequence, and grabbing "the first number
       after Grand Total" silently picks up Cost instead of Mkt Value,
       which previously caused the verify step itself to compare against
       the wrong ground truth and approve a wrong extraction.
    2. Per-row plausibility -- value should never be wildly disproportionate
       to cost (catches the AI picking the wrong column on individual rows,
       which can partially cancel out in the total and slip past check #1).
    """
    if not funds:
        return {"ok": False, "reason": "no_funds_found"}

    total_extracted = sum(f.get("value", 0) for f in funds)
    if total_extracted <= 0:
        return {"ok": False, "reason": "zero_total_value"}

    # ── Check 1: reconcile against Grand Total's MKT VALUE column ──
    # Strategy: find the "Grand Total" line, then look at ALL numbers on
    # that line/nearby text. The Mkt Value is the 2nd large rupee figure
    # in sequence (Cost, Mkt Value, Unrealized G/L, ...) -- not simply
    # the first number found after the words "Grand Total".
    stated_total = None
    grand_total_match = re.search(r'grand\s*total([^\n]{0,200})', raw_text, re.I)
    if grand_total_match:
        line_after = grand_total_match.group(1)
        nums = [float(n.replace(",", "")) for n in re.findall(r'[\d,]{4,}\.?\d*', line_after)]
        # Filter to plausible portfolio-value-sized numbers (reject tiny
        # ones like XIRR% or holding-period-months that might get caught)
        large_nums = [n for n in nums if n >= 1000]
        if len(large_nums) >= 2:
            # Column order is Cost, Mkt Value, Unrealized G/L, ... --
            # so index 1 (the SECOND large number) is Mkt Value
            stated_total = large_nums[1]
        elif len(large_nums) == 1:
            stated_total = large_nums[0]

    # Also try alternative explicit labels some platforms use
    if stated_total is None:
        for pat in [r'total\s*portfolio\s*value[^\d]{0,30}?([\d,]{4,}\.?\d*)',
                    r'total\s*current\s*value[^\d]{0,30}?([\d,]{4,}\.?\d*)']:
            m = re.search(pat, raw_text, re.I)
            if m:
                try:
                    stated_total = float(m.group(1).replace(",", ""))
                    break
                except ValueError:
                    continue

    # ── Check 2: per-row plausibility (value vs cost ratio) ──
    # Tightened from 10x to 5.5x after a real extraction error slipped
    # through at ~6x (Franklin India Mid Cap folio mismerge found in
    # testing) -- a 5x+ gain (500%) on a single holding is rare enough in
    # practice that it's worth a closer look rather than auto-accepting.
    implausible_rows = []
    for f in funds:
        cost = f.get("cost", 0) or 0
        value = f.get("value", 0) or 0
        if cost > 0 and value > 0:
            ratio = value / cost
            if ratio > 5.5 or ratio < 0.1:
                implausible_rows.append({"name": f.get("name"), "cost": cost, "value": value, "ratio": round(ratio, 2)})

    if stated_total is None:
        if implausible_rows:
            return {"ok": False, "reason": "implausible_rows", "extracted_total": total_extracted,
                    "implausible_rows": implausible_rows[:5]}
        return {"ok": True, "reason": "no_total_to_check", "extracted_total": total_extracted}

    diff_pct = abs(total_extracted - stated_total) / stated_total * 100
    if diff_pct <= 5 and not implausible_rows:
        return {"ok": True, "reason": "reconciled", "extracted_total": total_extracted, "stated_total": stated_total}
    return {"ok": False, "reason": "totals_mismatch" if diff_pct > 5 else "implausible_rows",
            "extracted_total": total_extracted, "stated_total": stated_total,
            "diff_pct": round(diff_pct, 1), "implausible_rows": implausible_rows[:5]}


async def _agent_extract_cas_via_ai(api_key: str, raw_text: str) -> list:
    """REPAIR/PRIMARY step -- ask Claude to read the raw statement text
    directly and extract every fund holding, regardless of provider-
    specific layout. This replaces needing a new hand-coded parser for
    every new statement format/provider encountered.

    Long statements (multi-page, 50+ holdings) are chunked rather than
    truncated -- a hard truncation silently drops holdings from large
    portfolios, which is worse than a slower multi-call extraction.
    """
    if not api_key:
        return []

    # Chunk size tuned to stay well under context/output limits while
    # keeping each chunk's table rows intact (a holding row is rarely
    # split across a chunk boundary at this size in practice; any
    # split row is caught by the VERIFY reconciliation check afterward)
    CHUNK_SIZE = 12000
    chunks = [raw_text[i:i+CHUNK_SIZE] for i in range(0, len(raw_text), CHUNK_SIZE)] or [raw_text]

    extraction_instructions = (
        f'This is raw text extracted from a client mutual fund portfolio statement '
        f'(could be from CAMS, KFintech, or a platform like Scripbox, Groww, Kuvera, '
        f'ETMoney, INDmoney -- the exact layout varies by provider).\n\n'
        f'Extract EVERY individual mutual fund HOLDING row mentioned -- each row '
        f'represents one folio/purchase.\n\n'
        f'COLUMN MAPPING -- these statements typically have MANY numeric columns per '
        f'row (Cost, Mkt Value, Unrealized G/L, Realized Gain/Loss, Dividend Paid '
        f'Since Inception, XIRR %, Total Returns, % to Portfolio, Holding Period, '
        f'Purchase Price per unit, Nav per unit, G/L %). You must map them precisely '
        f'-- do NOT guess or substitute one for another:\n'
        f'- "value" = the column literally labeled "Mkt Value" (current market value '
        f'of the holding TODAY). This is usually the 2nd large rupee figure on the '
        f'row, right after Cost. It is NOT "Total Returns", NOT "Dividend Paid Since '
        f'Inception", and NOT the per-unit "Nav" price.\n'
        f'- "cost" = the column literally labeled "Cost" or "Purchase Price" total '
        f'(the original investment amount). This is usually the 1st large rupee '
        f'figure on the row, before Mkt Value.\n'
        f'- "units" = the units held (a decimal number, typically 1-6 digits before '
        f'the decimal point, e.g. 4808.151) -- NOT the folio number, NOT the NAV.\n'
        f'- Ignore entirely: Unrealized G/L, Realized Gain/Loss, Dividend Paid Since '
        f'Inception, XIRR %, Total Returns, % to Portfolio, Holding Period (Months), '
        f'per-unit Purchase Price, per-unit Nav, G/L %. These are NOT cost or value.\n\n'
        f'SANITY CHECK before including a row: cost + (any reasonable gain or loss) '
        f'should be in the same order of magnitude as value. If your extracted value '
        f'looks wildly different from cost (e.g. 10x+ larger with no clear gain '
        f'column supporting that), you have likely picked the wrong column -- re-read '
        f'that row carefully.\n\n'
        f'WATCH FOR REPEATED NUMBERS ACROSS FOLIOS OF THE SAME SCHEME: when one scheme '
        f'has multiple folios, several columns (XIRR %, "Total Returns", per-unit Nav) '
        f'are IDENTICAL across all of that scheme\'s rows -- this is normal, since '
        f'those reflect the scheme overall, not the individual folio. Do NOT let this '
        f'repetition cause you to merge two folio rows into one, split one row into '
        f'two, or copy a number from the wrong adjacent row. Each folio still has its '
        f'OWN distinct units, cost, and Mkt Value even when other columns repeat -- '
        f'read each folio\'s units/cost/value directly from that folio\'s own row, '
        f'never inferred from a neighboring row.\n\n'
        f'CRITICAL: The SAME scheme name can legitimately appear MULTIPLE TIMES with '
        f'different folio numbers (e.g. a client invested in "XYZ Mid Cap Fund" via '
        f'5 separate folios over the years). Each occurrence is a SEPARATE real '
        f'holding with its own value -- you must extract ALL of them, never merge or '
        f'deduplicate by scheme name. Include the folio number in your output so '
        f'each holding can be told apart.\n\n'
        f'Ignore category/subtotal rows (e.g. "Equity", "Large Cap", "Debt", '
        f'"Hybrid") and the final grand total row -- only extract individual fund '
        f'holding rows that have a folio number and a purchase date.\n\n'
        f'Return ONLY a JSON array, no other text, no markdown fences:\n'
        f'[{{"name":"exact fund scheme name","folio":"folio number as text","value":0.0,"cost":0.0,"units":0.0}}, ...]\n\n'
        f'Statement text (this may be one section of a longer statement -- extract '
        f'only what appears in THIS text, do not guess at content outside it):\n'
    )

    all_funds = []
    try:
        async with httpx.AsyncClient(timeout=90.0) as client:
            for chunk_idx, chunk in enumerate(chunks):
                prompt = extraction_instructions + chunk
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 4000,
                        "messages": [{"role": "user", "content": prompt}]
                    })
                if resp.status_code != 200:
                    log.warning(f"CAS AI extraction chunk {chunk_idx} HTTP {resp.status_code}")
                    continue
                text = resp.json().get("content", [{}])[0].get("text", "").strip()
                text = re.sub(r"```[^\n]*\n?|```", "", text).strip()
                start, end = text.find("["), text.rfind("]") + 1
                if start < 0 or end <= start:
                    continue
                try:
                    funds = json.loads(text[start:end])
                    valid = [f for f in funds if isinstance(f, dict) and f.get("name") and f.get("value", 0) >= 0]
                    all_funds.extend(valid)
                except json.JSONDecodeError as e:
                    log.warning(f"CAS AI extraction chunk {chunk_idx} JSON parse failed: {e}")
                    continue
    except Exception as e:
        log.warning(f"CAS AI extraction failed: {e}")

    # Dedupe ONLY on the exact (name, folio) pair -- never on name alone,
    # since the same scheme name across different folios are genuinely
    # separate holdings (this was the root cause of a real data-corruption
    # bug found in testing -- deduping by name alone silently discarded
    # real holdings and undercounted a client's portfolio by ~35%)
    seen = {}
    for f in all_funds:
        key = f"{f.get('name','').strip().lower()}__{f.get('folio','').strip()}"
        if key not in seen or f.get("value", 0) > seen[key].get("value", 0):
            seen[key] = f
    return list(seen.values())


# ====================================================================
# FACTSHEET BOOKLET EXTRACTION AGENT
#
# Handles AMC marketing factsheet booklets (NOT the SEBI portfolio
# disclosure format) -- a single PDF covering 60+ schemes, where each
# scheme shows only a Top-N holdings table (no ISIN, partial coverage)
# buried inside pages of marketing prose, performance tables, and fund
# manager bios. This is a fundamentally different document shape from
# the CAS/AMC-disclosure parsers built elsewhere in this file, and
# needed its own dedicated approach:
#
#   1. PLAN   -- the booklet's own INDEX page lists every scheme name
#                with its starting page number. Use that directly
#                instead of guessing where each scheme's content is.
#   2. ACT    -- for each scheme, pull the text of its first page (the
#                holdings table is reliably on the first page of each
#                scheme's 2-3 page spread) and ask AI to extract just
#                the Top Holdings table, explicitly distinguishing it
#                from other similarly-shaped tables on the same page
#                (e.g. "Top Contributors" which has different columns
#                but can superficially look similar).
#   3. VERIFY -- flag schemes where no holdings table was found, or
#                where percentages don't sum to a plausible range.
#   4. REPORT -- per-scheme extraction status, so gaps are visible
#                rather than silently missing.
# ====================================================================

def _parse_factsheet_index(doc) -> dict:
    """Parse the booklet's own INDEX page(s) into {scheme_name: page_index}.
    Scans the first 6 pages looking for the index (handles the index
    spanning more than one page, as Edelweiss's does)."""
    import re as _re
    SECTION_HEADERS = {
        'expert speaks', 'equity funds', 'hybrid funds', 'precious metals',
        'debt funds', 'equity and hybrid passive funds', 'debt passive funds',
        'debt fund of funds', 'overseas fund of funds', 'other details',
        'index', 'i n d e x',
    }
    index_text = ""
    for i in range(min(6, len(doc))):
        t = doc[i].get_text()
        if _re.search(r'i\s*n\s*d\s*e\s*x', t, _re.I) or index_text:
            index_text += "\n" + t
        if index_text and not _re.search(r'i\s*n\s*d\s*e\s*x', t, _re.I) and i > 2:
            # stop once we've moved past index pages (heuristic: a page
            # with long paragraphs rather than name+number pairs)
            words_per_line = [len(l.split()) for l in t.split('\n') if l.strip()]
            if words_per_line and sum(w > 8 for w in words_per_line) > len(words_per_line) * 0.5:
                break

    lines = [l.strip() for l in index_text.split('\n') if l.strip()]
    scheme_pages = {}
    i = 0
    while i < len(lines) - 1:
        line, nxt = lines[i], lines[i + 1]
        if line.lower() in SECTION_HEADERS or _re.match(r'i\s*n\s*d\s*e\s*x', line, _re.I):
            i += 1
            continue
        if _re.match(r'^\d+$', nxt) and len(line) > 5:
            scheme_pages[line] = int(nxt) - 1  # convert to 0-based PDF index
            i += 2
        else:
            i += 1
    return scheme_pages


async def _agent_extract_factsheet_scheme(api_key: str, scheme_name: str, page_text: str) -> dict:
    """Extract the Top Holdings table for ONE scheme from its factsheet
    page text. Returns {"holdings": [...], "status": "ok"|"not_found"}.
    """
    if not api_key:
        return {"holdings": [], "status": "no_api_key"}

    prompt = (
        f'This text is from one page of an AMC factsheet booklet, for the scheme '
        f'"{scheme_name}".\n\n'
        f'Find the EQUITY/PORTFOLIO HOLDINGS table -- it is usually titled "Top N '
        f'Holdings" or similar, with columns like Company Name, Allocation/Exposure %, '
        f'and sometimes a Domestic/International tag.\n\n'
        f'IMPORTANT -- do NOT confuse it with these OTHER tables that may appear on '
        f'the same page and look superficially similar:\n'
        f'- "Top Contributors" or "Top 10 Contributors" tables (have Weights % AND a '
        f'separate Contribution % column -- this is a performance attribution table, '
        f'NOT the holdings list)\n'
        f'- "Portfolio Changes" (New Entries / Exits) -- just a list of names, no '
        f'percentages, not the holdings table\n'
        f'- Sector exposure or Market Capitalization tables (categories like "Large '
        f'Cap", "IT", "Banks" with percentages -- these are NOT individual stock '
        f'holdings)\n\n'
        f'Extract ONLY the actual stock-by-stock holdings table. For each holding, '
        f'capture the company name, its allocation percentage, and the Domestic/'
        f'International tag if present (default to "Domestic" if no tag is shown).\n\n'
        f'Return ONLY JSON, no other text:\n'
        f'{{"holdings":[{{"name":"company name","pct":0.0,"exposure":"Domestic|International"}}, ...],'
        f'"status":"ok"}}\n'
        f'If you cannot find any holdings table on this page, return: '
        f'{{"holdings":[],"status":"not_found"}}\n\n'
        f'Page text:\n{page_text[:6000]}'
    )
    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 3000,
                    "messages": [{"role": "user", "content": prompt}]
                })
            if resp.status_code != 200:
                log.warning(f"Factsheet extraction HTTP {resp.status_code} for '{scheme_name}'")
                return {"holdings": [], "status": f"api_error_{resp.status_code}"}
            text = resp.json().get("content", [{}])[0].get("text", "").strip()
            text = re.sub(r"```[^\n]*\n?|```", "", text).strip()
            start, end = text.find("{"), text.rfind("}") + 1
            if start < 0 or end <= start:
                return {"holdings": [], "status": "no_json_in_response"}
            result = json.loads(text[start:end])
            holdings = [h for h in result.get("holdings", [])
                        if isinstance(h, dict) and h.get("name") and h.get("pct", 0) > 0]
            return {"holdings": holdings, "status": result.get("status", "ok") if holdings else "not_found"}
    except Exception as e:
        log.warning(f"Factsheet extraction failed for '{scheme_name}': {e}")
        return {"holdings": [], "status": f"exception_{str(e)[:50]}"}


@app.post("/factsheet-extract")
async def factsheet_extract(payload: dict):
    """Extract Top-N holdings for every scheme found in an AMC factsheet
    booklet PDF. Uses the booklet's own INDEX page to find each scheme's
    page reliably, then runs AI extraction per scheme page.

    Request body: {"pages_text": ["page0 text", "page1 text", ...], "amc": "..."}
    (Client extracts text per page client-side via PDF.js and sends the
    full array -- avoids needing a PDF library server-side.)

    Response: {"schemes": {"Scheme Name": {"holdings":[...], "status":"ok"}, ...},
               "index_found": N, "extracted_ok": N, "not_found": [...]}
    """
    pages_text = payload.get("pages_text") or []
    amc_name = payload.get("amc", "")
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")

    if not pages_text:
        return {"error": "pages_text is required"}

    # PLAN -- parse the index from the first few pages
    import re as _re
    SECTION_HEADERS = {
        'expert speaks', 'equity funds', 'hybrid funds', 'precious metals',
        'debt funds', 'equity and hybrid passive funds', 'debt passive funds',
        'debt fund of funds', 'overseas fund of funds', 'other details',
        'index', 'i n d e x',
    }
    index_text = "\n".join(pages_text[:6])
    lines = [l.strip() for l in index_text.split('\n') if l.strip()]
    scheme_pages = {}
    i = 0
    while i < len(lines) - 1:
        line, nxt = lines[i], lines[i + 1]
        if line.lower() in SECTION_HEADERS or _re.match(r'i\s*n\s*d\s*e\s*x', line, _re.I):
            i += 1
            continue
        if _re.match(r'^\d+$', nxt) and len(line) > 5:
            page_idx = int(nxt) - 1
            if 0 <= page_idx < len(pages_text):
                scheme_pages[line] = page_idx
            i += 2
        else:
            i += 1

    if not scheme_pages:
        return {"error": "Could not parse an index from this document",
                 "schemes": {}, "index_found": 0}

    # ACT -- extract holdings for each scheme from its index-pointed page
    schemes_result = {}
    not_found = []
    for scheme_name, page_idx in scheme_pages.items():
        page_text = pages_text[page_idx]
        result = await _agent_extract_factsheet_scheme(api_key, scheme_name, page_text)
        schemes_result[scheme_name] = result
        if result["status"] != "ok" or not result["holdings"]:
            not_found.append(scheme_name)

    extracted_ok = len(scheme_pages) - len(not_found)
    log.info(f"Factsheet agent [{amc_name}]: {extracted_ok}/{len(scheme_pages)} schemes extracted, "
             f"{len(not_found)} not found")

    return {
        "schemes": schemes_result,
        "index_found": len(scheme_pages),
        "extracted_ok": extracted_ok,
        "not_found": not_found,
    }


@app.post("/cas-extract")
async def cas_extract(payload: dict):
    """CAS Extraction Agent -- decides between a fast free deterministic
    parse and AI extraction based on statement complexity, then verifies
    the result reconciles against the document's own stated total.

    Testing against real multi-page client statements (50-100+ holdings,
    repeated scheme names across multiple folios, segregated/zero-value
    portfolios) showed the regex-based deterministic parser is reliable
    only for simple statements -- on complex ones it silently miscounts
    by 30%+ (e.g. by mismatching which numbers belong to which row, or
    by a dedup step incorrectly merging genuinely separate folios of the
    same scheme). Rather than keep patching the regex per failure mode,
    AI extraction is now the PRIMARY path for any statement with more
    than a handful of holdings -- the regex stays only as an instant,
    free path for trivial 1-4 fund statements where it's reliable and
    saves an API call.

    Request body: {"raw_text": "...", "parsed_funds": [{"name","value","cost","units"}, ...]}
    Response: {"funds": [...], "source": "deterministic"|"ai_primary"|"ai_repair", "verify": {...}}
    """
    raw_text = payload.get("raw_text", "")
    parsed_funds = payload.get("parsed_funds") or []
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")

    SIMPLE_STATEMENT_THRESHOLD = 5  # at or below this many detected holdings, trust the regex

    if len(parsed_funds) > SIMPLE_STATEMENT_THRESHOLD and api_key:
        # AI-FIRST PATH -- statement is complex enough that the deterministic
        # parser's failure modes (column misalignment, incorrect dedup) are
        # a real risk. Go straight to AI extraction rather than trusting a
        # regex result that might already be silently wrong.
        ai_funds = await _agent_extract_cas_via_ai(api_key, raw_text)
        if ai_funds:
            ai_verdict = _cas_verify_extraction(ai_funds, raw_text)
            log.info(f"CAS agent (AI-primary, {len(parsed_funds)} detected holdings): "
                     f"AI found {len(ai_funds)} funds, reconciled={ai_verdict['ok']}")
            return {"funds": ai_funds, "source": "ai_primary", "verify": ai_verdict}
        # AI extraction failed entirely (no API key issue, network, etc) --
        # fall back to the deterministic result with a clear low-confidence flag
        verdict = _cas_verify_extraction(parsed_funds, raw_text)
        log.warning("CAS agent: AI-primary path failed, falling back to deterministic result")
        return {"funds": parsed_funds, "source": "deterministic_fallback", "verify": verdict}

    # SIMPLE STATEMENT -- trust the free deterministic parse, but still verify
    verdict = _cas_verify_extraction(parsed_funds, raw_text)
    if verdict["ok"]:
        return {"funds": parsed_funds, "source": "deterministic", "verify": verdict}

    # REPAIR -- even a "simple" statement's deterministic parse failed
    # verification, fall back to AI
    ai_funds = await _agent_extract_cas_via_ai(api_key, raw_text)
    if ai_funds:
        ai_verdict = _cas_verify_extraction(ai_funds, raw_text)
        log.info(f"CAS agent: deterministic parse failed ({verdict['reason']}), "
                 f"AI repair found {len(ai_funds)} funds, reconciled={ai_verdict['ok']}")
        return {"funds": ai_funds, "source": "ai_repair", "verify": ai_verdict,
                "deterministic_failure_reason": verdict["reason"]}

    log.warning(f"CAS agent: both deterministic and AI extraction failed/empty. "
                f"Deterministic reason: {verdict['reason']}")
    return {"funds": parsed_funds, "source": "failed", "verify": verdict}


@app.post("/cas-resolve")
async def cas_resolve(payload: dict):
    """CAS Import Agent Pass 3 -- given a list of fund names that Pass 1
    (local) and Pass 2 (server /search threshold) both failed to match,
    use AI to make a final judgment call per fund. Designed to be called
    once per CAS import with the batch of still-unmatched funds.

    Request body: {"unmatched": ["fund name 1", "fund name 2", ...]}
    Response: {"resolutions": {"fund name 1": {...}, ...}}
    """
    unmatched = payload.get("unmatched", [])
    if not unmatched:
        return {"resolutions": {}}
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    resolutions = {}
    MAX_PER_CALL = 10  # bound cost per CAS import

    for cas_name in unmatched[:MAX_PER_CALL]:
        qn = norm(cas_name)
        qw = set(qn.split())
        candidates = []
        for k, v in holdings_db.items():
            kw = set(k.split())
            s = len(qw & kw) / max(min(len(qw), len(kw)), 1)
            if qn in k: s += 0.6
            if k in qn: s += 0.4
            if s > 0:
                candidates.append({"score": round(s, 2), "name": v["fund_name"],
                                    "amc": v["amc"], "key": k})
        candidates.sort(key=lambda x: -x["score"])

        resolutions[cas_name] = await _agent_resolve_fund_name(api_key, cas_name, candidates)

    return {"resolutions": resolutions, "cache_size": len(_cas_resolve_cache)}

@app.get("/amc-health")
async def amc_health():
    """Visible health log -- last N AMC upload cycles with per-fund verify/repair detail."""
    return {"uploads_logged": len(_amc_health), "history": _amc_health,
            "cap_cache_size": len(_cap_resolution_cache)}

@app.post("/amc-health/recheck")
async def amc_health_recheck(amc: str = Query(...), secret: str = ""):
    """Manually re-run the verify/repair agent against an AMC already in
    holdings_db, without needing to re-upload the file. Useful for backfilling
    health checks on AMCs uploaded before this agent existed."""
    check_secret(secret)
    matching = {k: v for k, v in holdings_db.items()
                if v.get("amc", "").lower() == amc.lower()}
    if not matching:
        raise HTTPException(404, f"No funds found for AMC '{amc}'")
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    report = await _run_amc_upload_agent(amc, matching, api_key)
    return report

def norm(s: str) -> str:
    n = str(s).lower().strip()
    # Strip regulatory parenthetical suffix: "(Mid Cap Fund- An open ended...)"
    n = re.sub(r'\s*\([^)]{20,}\).*$', '', n)  # remove long parentheticals
    n = re.sub(r'\s*-?\s*(direct|regular)\s*plan.*$', '', n, flags=re.I)
    n = re.sub(r'\s*-?\s*(growth|idcw|dividend)\s*$', '', n, flags=re.I)
    n = re.sub(r'\s*\(g\)\s*$|\s*\(d\)\s*$', '', n, flags=re.I)
    # Expand common abbreviations
    n = re.sub(r'\bpru\b', 'prudential', n)
    n = re.sub(r'\bfof\b', 'fund of funds', n)
    n = re.sub(r'\betf\b', 'etf', n)
    return re.sub(r'\s+', ' ', n).strip()

VALID_ISIN = re.compile(r'^IN[A-Z0-9]{10}$')

SKIP_ROW = re.compile(
    r'^(sub.?total|grand.?total|total$|total\s+for|'
    r'equity\s*&|equity\s+and|debt\s+instru|listed\s*/|unlisted|'
    r'money\s*market|government\s*securities$|'
    r'treps?|reverse\s*repo|margin|cblo|'
    r'net\s*current|cash\s*and\s*other|'
    r'a\)|b\)|c\)|d\)|e\)|\(a\)|\(b\)|\(c\)|\(d\)|'
    r'nil$|n\.a\.$)',
    re.I
)

CASH_ROW = re.compile(
    r'^(treps?|triparty\s*repo|reverse\s*repo|cblo|'
    r'cash\s*and\s*other\s*net|net\s*current\s*asset)',
    re.I
)


# =============================================================================
# CORE: Universal sheet parser
# =============================================================================
def parse_sheet_universal(rows: list, fund_name: str = "") -> tuple:
    """
    Parse rows into (holdings, cash_pct, fund_name).
    Auto-detects: ISIN col, name col, pct col, decimal vs actual %.
    """
    header_row_idx = isin_col = name_col = pct_col = -1

    for r, row in enumerate(rows[:15]):
        if not row: continue
        vals = {i: str(c or '').strip() for i, c in enumerate(row)}
        vl   = {i: v.lower() for i, v in vals.items()}

        # Pick up fund name from content before header
        if not fund_name and header_row_idx < 0:
            for i, v in vals.items():
                vlo = v.lower()
                if vlo.startswith('scheme:') and len(v) > 10:
                    fund_name = v[7:].strip(); break
                if vlo.startswith('portfolio of '):
                    m = re.match(r'portfolio of (.+?)\s+as on', v, re.I)
                    if m: fund_name = m.group(1).strip(); break
                if re.search(r'scheme\s*name', vlo) and i + 1 in vals:
                    nv = vals.get(i + 1, '')
                    if len(nv) > 5 and 'scheme' not in nv.lower():
                        fund_name = nv; break
                # Motilal: R1 = ['Back to Index', 'Fund Name', ...]
                if vlo in ('back to index', 'back to table of contents') and i + 1 in vals:
                    nv = vals.get(i + 1, '')
                    if len(nv) > 8 and any(k in nv.lower() for k in ['fund', 'etf', 'scheme']):
                        fund_name = nv; break
                if (i < 3 and len(v) > 8 and 'fund' in vlo
                        and 'mutual fund' not in vlo
                        and 'asset management' not in vlo
                        and not fund_name):
                    # Strip the long SEBI-mandated scheme description AMCs
                    # append in parentheses (e.g. "HDFC Low Duration Fund
                    # (An open ended low duration debt scheme investing in
                    # instruments such that...)") -- this is regulatory
                    # boilerplate, not part of the fund's actual name, and
                    # was previously shown in full across the UI everywhere
                    # this detected name surfaces (fund cards, dropdowns,
                    # etc). Fixed once here since this detection is shared
                    # by every caller (parse_single_fund, parse_kotak_style,
                    # parse_uti all route through this same function).
                    fund_name = re.sub(r'\s*\([^)]{20,}\)\s*$', '', v).strip()

        if any('isin' in v for v in vl.values()):
            header_row_idx = r
            sector_col_detected = -1
            for ci, v in vl.items():
                if 'isin' in v and isin_col < 0:
                    isin_col = ci
                if name_col < 0 and any(k in v for k in
                        ['name of', 'instrument', 'issuer', 'company/issuer', 'company']):
                    name_col = ci
                if pct_col < 0 and re.search(
                        r'%\s*(to|of|net)\s*(nav|aum|asset|net)|nav\s*%|aum\s*%'
                        r'|percentage\s*(to|of)\s*(nav|aum|asset|net)', v):
                    pct_col = ci
                # Detect the actual rating/sector/industry column by header
                # text rather than assuming a fixed offset from ISIN --
                # different AMCs put Coupon%, Quantity, or other columns
                # immediately after ISIN, with Rating/Industry further
                # along (e.g. HDFC: ISIN, Coupon%, Name, Industry+/Rating).
                # A fixed isin_col+1 offset silently grabs the wrong
                # column on these layouts and loses all rating data.
                if sector_col_detected < 0 and re.search(
                        r'industry|sector|rating', v):
                    sector_col_detected = ci
            break

    if header_row_idx < 0 or isin_col < 0:
        return [], 0.0, fund_name

    if name_col < 0:
        name_col = max(0, isin_col - 1)

    if pct_col < 0:
        hrow = {i: str(c or '').strip().lower()
                for i, c in enumerate(rows[header_row_idx] or [])}
        for ci in range(max(hrow.keys(), default=0), isin_col, -1):
            if '%' in hrow.get(ci, '') or 'nav' in hrow.get(ci, '') or 'percentage' in hrow.get(ci, ''):
                pct_col = ci; break
        if pct_col < 0:
            # Last resort -- previously defaulted straight to the rightmost
            # column, which silently grabbed the wrong field entirely on
            # files with extra trailing columns (Yield/YTC/Maturity Date
            # after the real percentage column). Prefer a column closer to
            # the ISIN/name columns over the absolute rightmost one.
            pct_col = isin_col + 4 if (isin_col + 4) in hrow else max(hrow.keys(), default=isin_col + 4)

    # Detect name offset (Kotak-style indented rows)
    actual_name_col = name_col
    for row in rows[header_row_idx + 1: header_row_idx + 8]:
        if not row: continue
        vals = {i: str(c or '').strip() for i, c in enumerate(row)}
        if VALID_ISIN.match(vals.get(isin_col, '')):
            if not vals.get(name_col, ''):
                for ci in range(isin_col - 1, -1, -1):
                    v = vals.get(ci, '')
                    if v and len(v) > 2 and not SKIP_ROW.match(v):
                        actual_name_col = ci; break
            break

    # Detect decimal % — sample up to 5 ISIN rows, use max value
    # If max pct value < 1.0, it's decimal (0.6035 = 60.35%)
    # If any value > 1.0, it's actual % (60.35 = 60.35%)
    pct_is_decimal = False
    pct_samples = []
    for row in rows[header_row_idx + 1: header_row_idx + 20]:
        if not row: continue
        vals = {i: str(c or '').strip() for i, c in enumerate(row)}
        if VALID_ISIN.match(vals.get(isin_col, '')):
            raw = vals.get(pct_col, '').replace('%', '').replace(',', '').replace('$', '').strip()
            raw = re.sub(r'[^\d.\-]', '', raw)
            try:
                pv = float(raw)
                if pv > 0: pct_samples.append(pv)
            except:
                pass
        if len(pct_samples) >= 5: break
    if pct_samples:
        pct_is_decimal = max(pct_samples) < 1.0

    holdings  = []
    cash_pct  = 0.0

    for row in rows[header_row_idx + 1:]:
        if not row: continue
        vals = {i: str(c or '').strip() for i, c in enumerate(row)}

        isin_val = vals.get(isin_col, '')
        name_val = vals.get(actual_name_col, '').strip()

        if not VALID_ISIN.match(isin_val):
            if name_val and CASH_ROW.match(name_val):
                raw = vals.get(pct_col, '').replace('%', '').replace(',', '').strip()
                try:
                    p = float(raw)
                    if pct_is_decimal: p *= 100
                    if 0 < p < 50: cash_pct += p
                except:
                    pass
            continue

        if not name_val or len(name_val) < 2: continue
        if SKIP_ROW.match(name_val): continue

        # UTI-style prefixes ("NCD - ", "CD - ", "CP - ", "TB - " etc.) are a
        # strong, unambiguous debt-instrument signal -- NCD/CD/CP/TB/GB/DB are
        # never equity. Previously this prefix was stripped from the name
        # and silently discarded, leaving classification with no sector
        # text to go on and defaulting every such holding to "equity" --
        # a real bug that affected UTI-format debt funds (Low Duration,
        # Liquid, etc) where most/all holdings carry one of these prefixes.
        # Fix: capture the prefix's debt category into a synthetic sector
        # tag BEFORE stripping it from the display name, so the existing
        # DEBT_SECTOR_RE classification downstream has something to match.
        _prefix_m = re.match(r'^(EQ|DB|NCD|CP|TB|GB|MF|CB|CD|SPN/DDB|SPN|DDB|PTC|ZCB)\s*[-\u2013]\s*', name_val)
        prefix_debt_tag = ''
        if _prefix_m:
            _ptag = _prefix_m.group(1).upper()
            if _ptag != 'EQ':  # EQ explicitly means equity -- everything else here is debt
                prefix_debt_tag = {
                    'NCD': 'NCD', 'CP': 'Commercial Paper', 'TB': 'T-Bill',
                    'GB': 'G-Sec', 'DB': 'Debenture', 'CB': 'Bond', 'MF': 'MF',
                    'CD': 'Certificate of Deposit', 'SPN/DDB': 'Debt',
                    'SPN': 'Debt', 'DDB': 'Debt', 'PTC': 'Debt', 'ZCB': 'Debt',
                }.get(_ptag, _ptag)
        name_val = re.sub(r'^(EQ|DB|NCD|CP|TB|GB|MF|CB|CD|SPN/DDB|SPN|DDB|PTC|ZCB)\s*[-\u2013]\s*', '', name_val).strip()
        # Also catch debt-instrument signatures embedded in the name without
        # a clean "PREFIX - " pattern (e.g. "182 DAYS T-BILL - 30/04/2026",
        # "CD - KOTAK..." already caught above, but bare "T-BILL"/"CD"/"CP"
        # tokens elsewhere in the name are an equally strong signal)
        if not prefix_debt_tag and re.search(r'\bt-?bill\b|\bcertificate of deposit\b|\bcommercial paper\b|\bncd\b|\bdebenture\b|\bcorporate debt\b|\bdebt market\b|\bsecuritisation\b|\bsecuritization\b|\bptc\b', name_val, re.I):
            prefix_debt_tag = 'Debt'

        raw_pct = vals.get(pct_col, '').replace('%', '').replace(',', '').strip()
        raw_pct = re.sub(r'[^\d.\-]', '', raw_pct)
        try:
            pct = float(raw_pct)
        except:
            continue

        if pct_is_decimal:
            pct *= 100

        if pct <= 0 or pct > 100:
            continue

        # Use the header-detected rating/industry column if found; fall
        # back to the old isin_col+1 heuristic only when no such header
        # exists (some AMC layouts genuinely do put sector right after
        # ISIN with no explicit header text to detect).
        sector_col = sector_col_detected if sector_col_detected >= 0 else isin_col + 1
        sector = vals.get(sector_col, '')
        if sector and (re.match(r'^[\d.]+$', sector) or '%' in sector):
            sector = ''
        # If no real rating/sector was found in the data, but we captured
        # a debt-instrument prefix tag from the name (NCD/CP/TB/etc), use
        # that as the sector so downstream debt/equity classification has
        # a signal to work with instead of silently defaulting to equity.
        if not sector and prefix_debt_tag:
            sector = prefix_debt_tag

        holdings.append({
            "name":   name_val,
            "isin":   isin_val,
            "sector": sector,
            "pct":    round(pct, 4),
        })

    return holdings, round(cash_pct, 4), fund_name


# =============================================================================
# FORMAT DETECTORS
# =============================================================================
def _has_index_sheet(wb) -> bool:
    return any(s.strip().lower() == 'index' for s in wb.sheetnames)

def _is_uti_format(wb) -> bool:
    if len(wb.sheetnames) != 1: return False
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(max_row=3, values_only=True):
        if row and row[0] and 'SCHEME CODE' in str(row[0]).upper():
            return True
    return False

def _is_kotak_format(wb) -> bool:
    if _has_index_sheet(wb): return False
    for sname in wb.sheetnames[:3]:
        ws = wb[sname]
        rows = list(ws.iter_rows(max_row=2, values_only=True))
        if rows and len(rows[0]) > 2:
            cell = str(rows[0][2] or '').strip()
            if re.match(r'portfolio of .+ as on', cell, re.I):
                return True
    return False

def _is_single_fund(wb) -> bool:
    """Detects the common 'one main holdings sheet + supplementary sheets'
    pattern (e.g. HSBC: 'HEHYBF' + 'Notes' + 'Disclaimer'). Previously this
    required EXACTLY 1 sheet total, which incorrectly rejected files that
    have the single-fund layout but also include standard supplementary
    sheets like Notes/Disclaimer/Legend -- those are near-universal
    across AMC disclosures and don't make a file multi-fund.
    """
    SUPPLEMENTARY_SHEET_NAMES = {
        'notes', 'disclaimer', 'legend', 'abbreviations', 'glossary',
        'definitions', 'disclosure', 'index', 'cover', 'contents',
    }
    # Identify "real" data sheets -- those NOT matching a known
    # supplementary sheet name
    data_sheets = [s for s in wb.sheetnames if s.strip().lower() not in SUPPLEMENTARY_SHEET_NAMES]
    if len(data_sheets) != 1:
        return False
    ws = wb[data_sheets[0]]
    for row in ws.iter_rows(max_row=8, values_only=True):
        if any('isin' in str(c or '').lower() for c in row):
            return True
    return False


# =============================================================================
# FORMAT E: UTI
# =============================================================================
def parse_uti(wb, amc_name: str) -> dict:
    out = {}
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))

    current_fund = None
    fund_rows    = []

    def flush(fund_name, rows, out):
        if not fund_name or not rows: return
        holdings, cash_pct, _ = parse_sheet_universal(rows, fund_name)
        if len(holdings) >= 2:
            total = sum(h['pct'] for h in holdings)
            if total >= 5:
                key = norm(fund_name)
                out[key] = {"fund_name": fund_name, "amc": amc_name,
                            "holdings": holdings, "count": len(holdings),
                            "cashPct": cash_pct,
                            "uploaded_at": datetime.utcnow().isoformat(),
                            "format": "uti"}

    for row in all_rows:
        v0 = str(row[0] or '').strip() if row else ''
        vu = v0.upper()

        if 'SCHEME CODE' in vu and 'STARTS' in vu:
            flush(current_fund, fund_rows, out)
            current_fund = None; fund_rows = []; continue

        if 'SCHEME CODE' in vu and 'ENDS' in vu:
            flush(current_fund, fund_rows, out)
            current_fund = None; fund_rows = []; continue

        if v0.upper().startswith('SCHEME:'):
            current_fund = v0[7:].strip(); fund_rows = []; continue

        if current_fund:
            fund_rows.append(row)

    flush(current_fund, fund_rows, out)
    log.info(f"UTI: {len(out)} funds")
    return out


# =============================================================================
# FORMAT A/B: Multi-sheet with Index (Sundaram, Nippon, Axis, ABSL, SBI)
# =============================================================================
def parse_multi_sheet_with_index(wb, amc_name: str) -> dict:
    out = {}
    skip_sheets = {'index', 'cover', 'summary', 'disclaimer', 'annexure',
                   'contents', 'readme', 'back', 'annexure-a'}

    # Build index map: sheet_code → full fund name
    # Handles multiple layouts:
    #   Standard: [code, name] — Sundaram/Nippon/Axis/ABSL/SBI
    #   Motilal:  [serial, name, code] — code is last, name is middle
    index_map  = {}
    idx_sheet  = next((s for s in wb.sheetnames if s.strip().lower() == 'index'), None)
    if idx_sheet:
        for row in wb[idx_sheet].iter_rows(values_only=True):
            if not row: continue
            vals = [str(c or '').strip() for c in row]
            # Try standard pattern: scan for [code, name] pair
            found = False
            for i in range(len(vals) - 1):
                code = vals[i]; name = vals[i + 1] if i + 1 < len(vals) else ''
                if (re.match(r'^[A-Z0-9\-]{2,20}$', code)
                        and len(name) > 8
                        and any(k in name.lower() for k in ['fund', 'scheme', 'plan'])):
                    index_map[code] = name; found = True; break
            if not found:
                # Try Motilal pattern: [serial, name, code] — name before code
                for i in range(len(vals) - 1):
                    name = vals[i]; code = vals[i + 1] if i + 1 < len(vals) else ''
                    if (len(name) > 8
                            and any(k in name.lower() for k in ['fund', 'scheme', 'plan'])
                            and re.match(r'^[A-Z0-9\-]{2,20}$', code)):
                        index_map[code] = name; break

    for sname in wb.sheetnames:
        if sname.strip().lower() in skip_sheets: continue
        ws   = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4: continue

        fund_name = index_map.get(sname, '')
        holdings, cash_pct, detected = parse_sheet_universal(rows, fund_name)
        if not fund_name: fund_name = detected or sname

        if len(holdings) >= 2:
            total = sum(h['pct'] for h in holdings)
            if total < 5 or total > 200: continue
            key = norm(fund_name)
            out[key] = {"fund_name": fund_name.strip(), "amc": amc_name,
                        "holdings": holdings, "count": len(holdings),
                        "cashPct": cash_pct,
                        "uploaded_at": datetime.utcnow().isoformat(),
                        "format": "multi_index"}

    log.info(f"Multi-index: {len(out)} funds from {len(wb.sheetnames)} sheets")
    return out


# =============================================================================
# FORMAT C: Kotak — multi-sheet, no Index
# =============================================================================
def parse_kotak_style(wb, amc_name: str) -> dict:
    out = {}
    skip_sheets = {'index', 'cover', 'summary', 'disclaimer', 'annexure',
                   'contents', 'readme', 'back'}

    for sname in wb.sheetnames:
        if sname.strip().lower() in skip_sheets: continue
        ws   = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3: continue

        fund_name = ''
        if rows[0] and len(rows[0]) > 2:
            cell = str(rows[0][2] or '').strip()
            m = re.match(r'portfolio of (.+?)\s+as on', cell, re.I)
            if m: fund_name = m.group(1).strip()

        holdings, cash_pct, detected = parse_sheet_universal(rows, fund_name)
        if not fund_name: fund_name = detected or sname

        if len(holdings) >= 2:
            total = sum(h['pct'] for h in holdings)
            if total < 5 or total > 200: continue  # skip bad parses
            key = norm(fund_name)
            out[key] = {"fund_name": fund_name, "amc": amc_name,
                        "holdings": holdings, "count": len(holdings),
                        "cashPct": cash_pct,
                        "uploaded_at": datetime.utcnow().isoformat(),
                        "format": "kotak"}

    log.info(f"Kotak-style: {len(out)} funds")
    return out


# =============================================================================
# FORMAT D: Single fund per file (ICICI, HDFC)
# =============================================================================
def parse_single_fund(wb, amc_name: str, filename: str = "") -> dict:
    out = {}
    # Use the same supplementary-sheet-aware selection as _is_single_fund --
    # don't assume the data sheet is always sheetnames[0], since some AMCs
    # order supplementary sheets (Notes/Disclaimer/Legend) before the main
    # holdings sheet.
    SUPPLEMENTARY_SHEET_NAMES = {
        'notes', 'disclaimer', 'legend', 'abbreviations', 'glossary',
        'definitions', 'disclosure', 'index', 'cover', 'contents',
    }
    data_sheets = [s for s in wb.sheetnames if s.strip().lower() not in SUPPLEMENTARY_SHEET_NAMES]
    sheet_name = data_sheets[0] if data_sheets else wb.sheetnames[0]
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3: return out

    fund_name = ''
    for row in rows[:4]:
        for c in (row or []):
            v = str(c or '').strip()
            if len(v) > 8 and 'fund' in v.lower() and 'mutual fund' not in v.lower():
                fund_name = v; break
        if fund_name: break
    if not fund_name:
        fund_name = re.sub(r'\.xlsx?$', '', filename, flags=re.I).strip()
    # Description-stripping for fund_name detected here is handled
    # centrally inside parse_sheet_universal (shared by all callers).
    fund_name = re.sub(r'\s*\([^)]{20,}\)\s*$', '', fund_name).strip()

    holdings, cash_pct, detected = parse_sheet_universal(rows, fund_name)
    if not fund_name: fund_name = detected or filename

    if len(holdings) >= 2:
        total = sum(h['pct'] for h in holdings)
        if total >= 5:
            key = norm(fund_name)
            out[key] = {"fund_name": fund_name, "amc": amc_name,
                        "holdings": holdings, "count": len(holdings),
                        "cashPct": cash_pct,
                        "uploaded_at": datetime.utcnow().isoformat(),
                        "format": "single_fund"}
    return out


# =============================================================================
# .XLS LOADER (xlrd wrapper — Nippon, ABSL)
# =============================================================================
def load_xls_as_wb(raw: bytes):
    if not HAS_XLRD:
        log.warning(".xls file but xlrd not installed")
        return None
    try:
        wb_xls = xlrd.open_workbook(file_contents=raw)
    except Exception as e:
        log.warning(f"xlrd failed: {e}"); return None

    class FakeSheet:
        def __init__(self, ws): self._ws = ws
        def iter_rows(self, values_only=True, max_row=None):
            nrows = self._ws.nrows
            if max_row: nrows = min(nrows, max_row)
            for r in range(nrows):
                yield tuple(self._ws.cell_value(r, c) for c in range(self._ws.ncols))

    class FakeWB:
        def __init__(self, wb):
            self.sheetnames = wb.sheet_names()
            self._sheets = {n: FakeSheet(wb.sheet_by_name(n)) for n in self.sheetnames}
        def __getitem__(self, n): return self._sheets[n]
        def close(self): pass

    return FakeWB(wb_xls)


# =============================================================================
# ZIP EXTRACTOR
# =============================================================================
def extract_excels_from_zip(raw: bytes) -> list:
    results = []
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            for name in zf.namelist():
                if name.startswith(('__', '.')) or name.endswith('/'): continue
                data  = zf.read(name)
                fname = Path(name).name
                if fname.lower().endswith(('.xlsx', '.xls')):
                    results.append((fname, data))
                elif fname.lower().endswith('.zip'):
                    results.extend(extract_excels_from_zip(data))
    except Exception as e:
        log.warning(f"ZIP extract failed: {e}")
    return results


# =============================================================================
# MAIN DISPATCH
# =============================================================================
def open_workbook(raw: bytes, filename: str):
    if filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx'):
        wb = load_xls_as_wb(raw)
        if wb: return wb, 'xls'
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True, keep_links=False)
        return wb, 'xlsx'
    except Exception as e:
        log.warning(f"Cannot open '{filename}': {e}")
        return None, None


def process_upload(raw: bytes, filename: str, amc_name: str) -> dict:
    if filename.lower().endswith('.zip'):
        excels = extract_excels_from_zip(raw)
        log.info(f"ZIP '{filename}': {len(excels)} files")
        combined = {}
        for xname, xbytes in excels:
            combined.update(process_upload(xbytes, xname, amc_name))
        return combined

    wb, ftype = open_workbook(raw, filename)
    if not wb: return {}

    n = len(wb.sheetnames)
    log.info(f"'{filename}': {n} sheets [{ftype}]")

    if _is_uti_format(wb):
        log.info("  -> Format E (UTI)")
        result = parse_uti(wb, amc_name)
    elif _has_index_sheet(wb) and n > 3:
        log.info("  -> Format A/B (multi-sheet + Index)")
        result = parse_multi_sheet_with_index(wb, amc_name)
    elif _is_kotak_format(wb):
        log.info("  -> Format C (Kotak-style)")
        result = parse_kotak_style(wb, amc_name)
    elif _is_single_fund(wb) or n == 1:
        log.info("  -> Format D (single fund)")
        result = parse_single_fund(wb, amc_name, filename)
    else:
        log.info("  -> Fallback (multi-sheet no Index)")
        result = parse_kotak_style(wb, amc_name)

    try: wb.close()
    except: pass

    log.info(f"  -> {len(result)} funds extracted")
    return result


# =============================================================================
# AUTH + ROUTES
# =============================================================================
def check_secret(secret: str):
    exp = os.environ.get("UPLOAD_SECRET", "")
    if exp and secret != exp: raise HTTPException(403, "Invalid secret")

@app.get("/")
async def root():
    amcs = sorted({v["amc"] for v in holdings_db.values()})
    return {"service": "MF Holdings API v7", "funds": len(holdings_db), "amcs": amcs,
            "amfi_cap_loaded": len(AMFI_ISIN_CAP),
            "last_updated": max(
                (v.get("uploaded_at", "") for v in holdings_db.values()), default=None)}

@app.get("/health")
async def health():
    amcs = sorted({v["amc"] for v in holdings_db.values()})
    return {"status": "ok", "funds": len(holdings_db), "amcs": amcs}

@app.get("/debug-firestore")
async def debug_firestore():
    """Diagnostic: confirms whether Firestore persistence is actually
    configured and working, separate from whether holdings_db happens to
    have data in memory right now. Use this to verify the
    FIREBASE_SERVICE_ACCOUNT_JSON env var is set correctly on Render
    BEFORE relying on uploads surviving a restart.
    """
    sa_json_set = bool(os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", ""))
    db = _get_firestore_client()
    result = {
        "service_account_env_var_set": sa_json_set,
        "firestore_client_initialized": db is not None,
        "holdings_db_in_memory": len(holdings_db),
    }
    if db is not None:
        try:
            meta_ref = db.collection("mf_holdings_shards").document("_meta")
            meta = meta_ref.get()
            if meta.exists:
                result["firestore_meta"] = meta.to_dict()
            else:
                result["firestore_meta"] = "no _meta doc yet -- nothing saved to Firestore so far"
            # Round-trip write test -- proves the service account actually
            # has write permission, not just that the client initialized
            test_ref = db.collection("mf_holdings_shards").document("_healthcheck")
            test_ref.set({"ts": datetime.utcnow().isoformat()})
            test_read = test_ref.get()
            result["write_read_test"] = "PASSED" if test_read.exists else "FAILED"
        except Exception as e:
            result["firestore_error"] = str(e)[:300]
    else:
        result["reason"] = ("FIREBASE_SERVICE_ACCOUNT_JSON not set in Render's environment "
                            "variables" if not sa_json_set else
                            "Service account JSON is set but client initialization failed -- "
                            "check Render logs for the exact error")
    return result

@app.get("/funds")
async def list_funds(amc: Optional[str] = None):
    out = [{"name": v["fund_name"], "amc": v["amc"], "key": k,
            "count": v["count"], "uploaded_at": v.get("uploaded_at", ""),
            "format": v.get("format", "")}
           for k, v in holdings_db.items()
           if not amc or amc.lower() in v.get("amc", "").lower()]
    out.sort(key=lambda x: (x["amc"], x["name"]))
    return {"total": len(out), "funds": out}

@app.get("/holdings")
async def get_holdings(fund: str = Query(..., min_length=2)):
    if not holdings_db: raise HTTPException(503, "No data yet")
    q = norm(fund)
    if q in holdings_db:
        return _enrich_holdings(holdings_db[q])
    qw = set(q.split()); best_s, best_v = 0.0, None
    for k, v in holdings_db.items():
        kw = set(k.split())
        s  = len(qw & kw) / max(min(len(qw), len(kw)), 1)
        if q in k: s += 0.6
        if k in q: s += 0.4
        if s > best_s: best_s, best_v = s, v
    if best_v and best_s >= 0.4: return _enrich_holdings(best_v)
    raise HTTPException(404, f"Not found: '{fund}' (best={best_s:.2f})")

def _amfi_cap_map() -> dict:
    """Build name->cap dict from cached AMFI data."""
    d = _amfi_cap_cache.get("data", {})
    m = {}
    for cap in ("large", "mid", "small"):
        for name in d.get(cap, []):
            m[name] = cap
    return m

def _norm_stock(name: str) -> str:
    n = str(name).upper().strip()
    n = re.sub(r'\bLTD\.?\b|\bLIMITED\b|\bPVT\.?\b|\bPRIVATE\b', '', n)
    n = re.sub(r'[&]', 'AND', n)
    n = re.sub(r"['\"]", '', n)
    n = re.sub(r'[.\-,()\[\]]', ' ', n)
    return re.sub(r'\s+', ' ', n).strip()

DEBT_SECTOR_RE = re.compile(
    r'crisil|care|icra|fitch|ind-ra|aaa|aa\+|aa-|\baa\b|sovereign|'
    r'tbill|t-bill|treps|cblo|repo|gilt|g-sec|sdl|commercial paper|'
    r'certificate of deposit|fixed deposit|\bbond\b|debenture|ncd|'
    r'\bdebt\b|\bptc\b|securitisation|securitization', re.I)

# Equity sector patterns — if sector matches these, it's definitely equity
# even if DEBT_SECTOR_RE also matches (e.g. CRISIL Ltd in "Financial Services")
EQUITY_SECTOR_RE = re.compile(
    r'^(banks|insurance|it |software|pharma|auto|fmcg|consumer|capital goods|'
    r'industrial|financial services|healthcare|telecom|cement|metals|energy|'
    r'power|realty|media|retail|chemicals|textiles|agri|diversified)', re.I)

def _enrich_holdings(fund_data: dict) -> dict:
    """Add cap/type fields to each holding using server's AMFI data."""
    import copy
    cap_map = _amfi_cap_map()
    result = copy.copy(fund_data)
    enriched = []
    for h in fund_data.get("holdings", []):
        eh = dict(h)
        sector = h.get("sector", "")
        name   = h.get("name", "")
        # Classify instrument type
        # If sector == name, the parser found no sector — don't use it for debt detection
        effective_sector = "" if sector.strip() == name.strip() else sector
        is_debt = (bool(DEBT_SECTOR_RE.search(effective_sector)) and not bool(EQUITY_SECTOR_RE.match(effective_sector))) \
                  or bool(re.match(r'^\d+\.?\d*%', name))
        eh["type"] = "debt" if is_debt else "equity"
        # For equity: classify cap
        # Priority 1: ISIN lookup against bundled AMFI Excel (fast, accurate)
        # Priority 2: name-based lookup against live-fetched cache (fallback if file missing)
        # Priority 3: default 'small' — correct per SEBI (everything outside top 250 is small cap)
        if not is_debt:
            isin_val = h.get("isin", "").strip()
            cap = None
            if isin_val and AMFI_ISIN_CAP:
                cap = AMFI_ISIN_CAP.get(isin_val)
            if cap is None and cap_map:
                key = _norm_stock(name)
                cap = cap_map.get(key)
                if not cap and len(key) > 8:
                    pre = key[:12]
                    cap = next((cap_map[k] for k in cap_map if len(k)>8 and k[:12]==pre), None)
            eh["cap"] = cap or "small"
        else:
            eh["cap"] = None
        enriched.append(eh)
    result = dict(fund_data)
    result["holdings"] = enriched
    return result

@app.get("/search")
async def search(q: str = Query(..., min_length=2)):
    qn = norm(q); qw = set(qn.split()); res = []
    for k, v in holdings_db.items():
        kw = set(k.split())
        s  = len(qw & kw) / max(min(len(qw), len(kw)), 1)
        if qn in k: s += 0.6
        if k in qn: s += 0.4
        if s > 0:
            res.append({"score": round(s, 2), "name": v["fund_name"],
                        "amc": v["amc"], "key": k, "count": v["count"]})
    res.sort(key=lambda x: -x["score"])
    return {"query": q, "results": res[:15]}

@app.post("/upload-parsed")
async def upload_parsed(payload: dict):
    """Accept ALREADY-PARSED fund data (JSON) directly -- used by the
    client-side PDF.js factsheet parser for AMCs that only publish PDF
    disclosures (no Excel/ZIP available, e.g. Sundaram). Re-parsing a PDF
    server-side would duplicate work the client already did and risk
    inconsistent results between two separate parsers -- this endpoint
    instead trusts the client's PDF.js extraction and runs it through the
    EXACT SAME pipeline as Excel uploads: rename detection, holdings_db
    write, and the verify/repair Upload Agent.

    Expected payload shape:
    {
      "amc": "Sundaram Mutual Fund",
      "secret": "...",
      "funds": {
        "<normalized_key>": {
          "fund_name": "Sundaram Mid Cap Fund",
          "amc": "Sundaram Mutual Fund",
          "holdings": [{"name":"...","isin":"...","sector":"...","pct":1.23}, ...],
          "count": 47,
          "cashPct": 2.1
        }, ...
      }
    }
    """
    check_secret(payload.get("secret", ""))
    amc_name = (payload.get("amc") or "").strip()
    funds = payload.get("funds") or {}
    if not amc_name:
        raise HTTPException(422, "amc name is required")
    if not funds:
        raise HTTPException(422, "No fund data provided")

    # Validate and normalize each fund entry to the exact shape the rest
    # of the pipeline expects -- reject anything malformed rather than
    # silently accepting bad data into holdings_db
    parsed = {}
    for key, fund in funds.items():
        holdings = fund.get("holdings")
        if not isinstance(holdings, list) or len(holdings) < 2:
            continue  # same minimum-holdings bar as the Excel parser
        clean_holdings = []
        for h in holdings:
            if not isinstance(h, dict) or not h.get("name"):
                continue
            clean_holdings.append({
                "name":   str(h.get("name", "")).strip(),
                "isin":   str(h.get("isin", "")).strip(),
                "sector": str(h.get("sector", "")).strip(),
                "pct":    round(float(h.get("pct", 0) or 0), 4),
            })
        if len(clean_holdings) < 2:
            continue
        norm_key = norm(fund.get("fund_name", key))
        parsed[norm_key] = {
            "fund_name": str(fund.get("fund_name", key)).strip(),
            "amc": amc_name,
            "holdings": clean_holdings,
            "count": len(clean_holdings),
            "cashPct": round(float(fund.get("cashPct", 0) or 0), 4),
        }

    if not parsed:
        raise HTTPException(422, "No valid fund data after validation — check holdings format")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    rename_result = await _detect_and_resolve_renames(amc_name, parsed, api_key)

    holdings_db.update(parsed)
    save_db()

    asyncio.create_task(_run_amc_upload_agent(
        amc_name, parsed, api_key,
        rename_report={"renames_resolved": rename_result["renames_resolved"]}
    ))

    return {"status": "ok", "amc": amc_name, "source": "pdf_client_parsed",
            "funds_added": len(parsed), "funds_total": len(holdings_db),
            "renames_resolved": len(rename_result["renames_resolved"])}

@app.post("/upload")
async def upload(
    files:  List[UploadFile] = File(...),
    amc:    str              = Form(...),
    secret: str              = Form(default=""),
):
    """Upload AMC portfolio Excel/ZIP. Auto-detects format."""
    check_secret(secret)
    total_funds = 0; fund_names = []
    all_parsed = {}
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    all_renames_resolved = []

    for f in files:
        fname = f.filename or "upload"
        if not re.search(r'\.(xlsx|xls|zip)$', fname, re.I): continue
        raw    = await f.read()
        parsed = process_upload(raw, fname, amc.strip())
        if parsed:
            # Rename detection MUST run before holdings_db.update() -- once
            # the new key is written, we can no longer tell "new fund" apart
            # from "fund that already existed under this exact name".
            rename_result = await _detect_and_resolve_renames(amc.strip(), parsed, api_key)
            all_renames_resolved.extend(rename_result["renames_resolved"])

            holdings_db.update(parsed)
            all_parsed.update(parsed)
            total_funds += len(parsed)
            fund_names.extend(v["fund_name"] for v in parsed.values())
            log.info(f"[{amc}] '{fname}' -> {len(parsed)} funds")

    if total_funds == 0:
        raise HTTPException(422, "No fund data found — check file format")

    # Capture exactly what save_db() actually did -- surfaced directly in
    # this response so persistence failures are visible immediately
    # instead of requiring a separate /debug-firestore check after the
    # fact (which only shows the CURRENT state, not what happened during
    # THIS specific upload).
    save_status = {"local_cache": "unknown", "firestore": "unknown"}
    try:
        DB_FILE.write_text(json.dumps(holdings_db, ensure_ascii=False))
        save_status["local_cache"] = "ok"
    except Exception as e:
        save_status["local_cache"] = f"failed: {str(e)[:150]}"

    db = _get_firestore_client()
    if db is None:
        save_status["firestore"] = "client_unavailable (check FIREBASE_SERVICE_ACCOUNT_JSON)"
    else:
        try:
            keys = sorted(holdings_db.keys())
            shards = [keys[i:i + FIRESTORE_SHARD_SIZE] for i in range(0, len(keys), FIRESTORE_SHARD_SIZE)]
            batch = db.batch()
            for shard_idx, shard_keys in enumerate(shards):
                shard_data = {k: holdings_db[k] for k in shard_keys}
                doc_ref = db.collection("mf_holdings_shards").document(f"shard_{shard_idx}")
                batch.set(doc_ref, {"keys": shard_keys, "data": shard_data})
            batch.commit()
            db.collection("mf_holdings_shards").document("_meta").set({
                "shard_count": len(shards), "total_funds": len(holdings_db),
                "updated_at": datetime.utcnow().isoformat()
            })
            save_status["firestore"] = f"ok ({len(shards)} shards, {len(holdings_db)} total funds written)"
        except Exception as e:
            save_status["firestore"] = f"FAILED: {str(e)[:300]}"
            log.error(f"Firestore save failed during /upload: {e}")

    # Fire the verify/repair agent in the background — doesn't block the
    # upload response, but runs immediately so /amc-health reflects this
    # upload within a few seconds rather than requiring a manual trigger.
    if all_parsed:
        asyncio.create_task(_run_amc_upload_agent(
            amc.strip(), all_parsed, api_key,
            rename_report={"renames_resolved": all_renames_resolved}
        ))

    return {"status": "ok", "amc": amc, "files": len(files),
            "funds_added": total_funds, "funds_total": len(holdings_db),
            "funds": fund_names, "save_status": save_status}

@app.post("/preview")
async def preview_upload(
    files:  List[UploadFile] = File(...),
    amc:    str              = Form(...),
    secret: str              = Form(default=""),
):
    """Parse and preview without saving. Returns fund names + counts + total %."""
    check_secret(secret)
    preview = []
    for f in files:
        fname = f.filename or "upload"
        if not re.search(r'\.(xlsx|xls|zip)$', fname, re.I): continue
        raw    = await f.read()
        parsed = process_upload(raw, fname, amc.strip())
        for key, data in parsed.items():
            total_pct = sum(h['pct'] for h in data['holdings'])
            preview.append({
                "fund_name": data['fund_name'],
                "holdings":  data['count'],
                "total_pct": round(total_pct, 1),
                "format":    data.get('format', ''),
                "valid":     80 <= total_pct <= 115,
            })
    preview.sort(key=lambda x: x['fund_name'])
    valid = sum(1 for p in preview if p['valid'])
    return {"amc": amc, "total": len(preview),
            "valid": valid, "invalid": len(preview) - valid, "funds": preview}

@app.delete("/amc")
async def delete_amc(amc: str, secret: str = ""):
    check_secret(secret)
    keys = [k for k, v in holdings_db.items()
            if v.get("amc", "").lower() == amc.lower()]
    for k in keys: del holdings_db[k]
    save_db(); return {"deleted": len(keys), "funds_remaining": len(holdings_db)}

@app.delete("/fund")
async def delete_fund(key: str, secret: str = ""):
    check_secret(secret)
    if key not in holdings_db: raise HTTPException(404, f"Key not found: {key}")
    del holdings_db[key]; save_db()
    return {"deleted": key, "funds_remaining": len(holdings_db)}


# =============================================================================
# AMFI CAP PROXY
# =============================================================================
@app.get("/amfi-cap")
async def amfi_cap():
    import time, urllib.request
    global _amfi_cap_cache
    if _amfi_cap_cache and (time.time() - _amfi_cap_cache.get("ts", 0)) < 43200:
        return _amfi_cap_cache["data"]

    URLS = [
        ("Dec 2025", "https://www.amfiindia.com/Themes/Theme1/downloads/AverageMarketCapitalization31Dec2025.xlsx"),
        ("Jun 2025", "https://www.amfiindia.com/Themes/Theme1/downloads/AverageMarketCapitalization30Jun2025.xlsx"),
    ]

    def norm_name(n):
        n = str(n).upper().strip()
        n = re.sub(r'\bLTD\.?\b|\bLIMITED\b|\bPVT\.?\b', '', n)
        n = re.sub(r'[.\-,()]', ' ', n)
        return re.sub(r'\s+', ' ', n).strip()

    for label, url in URLS:
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.amfiindia.com/'})
            raw = urllib.request.urlopen(req, timeout=20).read()
            wb  = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws  = wb.active
            rows = list(ws.iter_rows(values_only=True))
            name_col = cat_col = rank_col = hdr_row = -1
            for i, row in enumerate(rows[:10]):
                cells = [str(c or '').lower().strip() for c in row]
                nc = next((j for j, c in enumerate(cells) if 'company' in c or 'name' in c), -1)
                cc = next((j for j, c in enumerate(cells) if 'categor' in c or 'large' in c), -1)
                rc = next((j for j, c in enumerate(cells)
                           if c in ('rank', 'sr', 'sr.', 'sl', 'no', 'no.')), -1)
                if nc >= 0: name_col=nc; cat_col=cc; rank_col=rc; hdr_row=i; break
            if name_col < 0: continue
            large, mid, small = [], [], []
            for row in rows[hdr_row + 1:]:
                if not row or not row[name_col]: continue
                name = norm_name(row[name_col])
                if not name or len(name) < 3: continue
                if cat_col >= 0:
                    cat = str(row[cat_col] or '').lower()
                    if 'large' in cat: large.append(name)
                    elif 'mid' in cat: mid.append(name)
                    elif 'small' in cat: small.append(name)
                elif rank_col >= 0:
                    try:
                        rank = int(row[rank_col])
                        if rank <= 100: large.append(name)
                        elif rank <= 250: mid.append(name)
                        else: small.append(name)
                    except: pass
            if len(large) >= 90:
                result = {"large": large, "mid": mid, "small": small,
                          "updated": label, "total": len(large)+len(mid)+len(small)}
                _amfi_cap_cache = {"ts": time.time(), "data": result}
                return result
        except Exception as e:
            log.warning(f"AMFI fetch failed ({label}): {e}")

    raise HTTPException(503, "AMFI data temporarily unavailable")


# =============================================================================
# MARKET MONITOR
# =============================================================================

_market_cache: dict = {"data": None, "ts": 0.0, "status": "idle"}
_indices_cache: dict = {"data": None, "ts": 0.0}
_news_cache: dict = {"data": None, "ts": 0.0}

NSE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/",
    "Origin": "https://www.nseindia.com",
}

async def _get_nse_session(client: httpx.AsyncClient):
    """Hit NSE homepage to get session cookies before API calls."""
    try:
        await client.get("https://www.nseindia.com", headers=NSE_HEADERS, timeout=10.0)
    except Exception:
        pass

YAHOO_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}

async def _fetch_indices() -> list:
    """Fetch live index data from Yahoo Finance (NSE direct scraping is blocked for
    non-Indian/datacenter IPs — Yahoo's unauthenticated chart API is a reliable
    free alternative that works from any server location)."""
    import time
    if _indices_cache["data"] and (time.time() - _indices_cache["ts"]) < 900:  # 15 min cache
        return _indices_cache["data"]
    results = []
    # Yahoo Finance ticker symbols for Indian indices
    index_map = [
        ("NIFTY 50",          "%5ENSEI"),
        ("NIFTY BANK",        "%5ENSEBANK"),
        ("NIFTY MIDCAP 150",  "NIFTY_MIDCAP_150.NS"),
        ("NIFTY SMALLCAP 250","NIFTY_SMLCAP_250.NS"),
        ("INDIA VIX",         "%5EINDIAVIX"),
        ("NIFTY IT",          "%5ECNXIT"),
    ]
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
            for label, symbol in index_map:
                try:
                    r = await client.get(
                        f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}",
                        params={"range": "1y", "interval": "1d"},
                        headers=YAHOO_HEADERS, timeout=8.0)
                    if r.status_code == 200:
                        d = r.json()
                        result = d.get("chart", {}).get("result")
                        if not result:
                            log.warning(f"Yahoo: no result for {label} ({symbol})")
                            continue
                        meta = result[0].get("meta", {})
                        last_price  = meta.get("regularMarketPrice", 0)
                        prev_close  = meta.get("chartPreviousClose") or meta.get("previousClose", 0)
                        change      = round(last_price - prev_close, 2) if last_price and prev_close else 0
                        change_pct  = round((change / prev_close) * 100, 2) if prev_close else 0
                        # 52-week high/low from the year of daily closes returned
                        closes = [c for c in (result[0].get("indicators",{}).get("quote",[{}])[0].get("close") or []) if c is not None]
                        year_high = max(closes) if closes else meta.get("fiftyTwoWeekHigh", 0)
                        year_low  = min(closes) if closes else meta.get("fiftyTwoWeekLow", 0)
                        results.append({
                            "name":      label,
                            "value":     round(last_price, 2),
                            "change":    change,
                            "changePct": change_pct,
                            "open":      meta.get("regularMarketOpen", 0) or 0,
                            "high":      meta.get("regularMarketDayHigh", 0) or 0,
                            "low":       meta.get("regularMarketDayLow", 0) or 0,
                            "yearHigh":  round(year_high, 2) if year_high else 0,
                            "yearLow":   round(year_low, 2) if year_low else 0,
                        })
                    else:
                        log.warning(f"Yahoo fetch failed for {label}: HTTP {r.status_code}")
                except Exception as e:
                    log.warning(f"Index fetch failed for {label}: {e}")
    except Exception as e:
        log.warning(f"Yahoo Finance session failed: {e}")
    if results:
        _indices_cache["data"] = results
        _indices_cache["ts"]   = time.time()
    return results

async def _fetch_fii_dii() -> dict:
    """Fetch FII/DII activity from NSE.
    NSE returns: [{category:'DII', date:'17-Apr-2026', buyValue:'17513.99',
                   sellValue:'22235.47', netValue:'-4721.48'},
                  {category:'FII/FPI', ...}]
    """
    def _flt(v):
        try: return float(str(v).replace(',', ''))
        except: return 0.0

    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
            await _get_nse_session(client)
            r = await client.get(
                "https://www.nseindia.com/api/fiidiiTradeReact",
                headers=NSE_HEADERS, timeout=8.0)
            if r.status_code != 200:
                log.warning(f"FII/DII HTTP {r.status_code}")
                return {}
            data = r.json()
            if not isinstance(data, list) or not data:
                return {}
            # Find FII and DII rows by category field
            fii = next((d for d in data if 'fii' in d.get('category','').lower() or 'fpi' in d.get('category','').lower()), None)
            dii = next((d for d in data if d.get('category','').lower() == 'dii'), None)
            if not fii or not dii:
                log.warning(f"FII/DII: could not find rows. Categories: {[d.get('category') for d in data]}")
                return {}
            date = fii.get('date') or dii.get('date') or 'Latest'
            result = {
                "date":          date,
                "fii_net_crore": _flt(fii.get('netValue', 0)),
                "dii_net_crore": _flt(dii.get('netValue', 0)),
                "fii_buy":       _flt(fii.get('buyValue', 0)),
                "fii_sell":      _flt(fii.get('sellValue', 0)),
                "dii_buy":       _flt(dii.get('buyValue', 0)),
                "dii_sell":      _flt(dii.get('sellValue', 0)),
            }
            log.info(f"FII/DII {date}: FII={result['fii_net_crore']}, DII={result['dii_net_crore']}")
            return result
    except Exception as e:
        log.warning(f"FII/DII fetch failed: {e}")
    return {}

# ════════════════════════════════════════════════════════════════
# MARKET DATA AGENT — Plan → Act → Verify → Repair → Report
#
# Replaces the old single-shot fetch with a self-checking loop:
#   1. PLAN   — decide what fields are required and how to search for them
#   2. ACT    — call Claude with web_search, run the searches
#   3. VERIFY — check the result against a schema: which fields are
#               present, missing, or implausible
#   4. REPAIR — if anything is missing, run ONE targeted follow-up
#               call asking specifically for the gaps (cheap — small
#               prompt, no full re-search of everything)
#   5. REPORT — every cycle is logged to _agent_health with what
#               happened, so failures are visible via /agent-health
#               instead of being discovered by chance
# ════════════════════════════════════════════════════════════════

REQUIRED_FIXED_INCOME_FIELDS = [
    "gsec_10y", "gsec_5y", "gsec_1y", "repo_rate", "rbi_stance",
    "cpi_inflation", "aaa_spread_10y", "mibor_overnight", "yield_curve_slope",
]

# Sensible recent fallback values — used only as an absolute last resort,
# always flagged to the client as estimated (never silently substituted).
FIXED_INCOME_FALLBACKS = {
    "gsec_10y": 6.85, "gsec_5y": 6.55, "gsec_1y": 6.45,
    "repo_rate": 5.25, "rbi_stance": "Neutral", "cpi_inflation": 3.2,
    "aaa_spread_10y": 55, "mibor_overnight": 5.3, "yield_curve_slope": 40,
}

_agent_health: list = []  # ring buffer of recent fetch-cycle reports, newest first
_AGENT_HEALTH_MAX = 20

def _agent_log(report: dict):
    """Record a fetch-cycle report into the rolling health log."""
    import time
    report["ts"] = time.time()
    report["time_str"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _agent_health.insert(0, report)
    while len(_agent_health) > _AGENT_HEALTH_MAX:
        _agent_health.pop()

def _agent_verify(result: dict) -> dict:
    """VERIFY step — inspect the result and report what's missing or implausible.
    Returns {"ok": bool, "missing_fixed_income": [...], "issues": [...]}.
    """
    issues = []
    fi = result.get("fixed_income") or {}
    missing = [f for f in REQUIRED_FIXED_INCOME_FIELDS if fi.get(f) in (None, "")]

    # Plausibility checks — catch cases where the model returned a number
    # that's technically present but nonsensical (e.g. a string, a wildly
    # out-of-range yield from a misread search result)
    for yield_field in ("gsec_10y", "gsec_5y", "gsec_1y", "repo_rate", "mibor_overnight"):
        v = fi.get(yield_field)
        if v is not None and not (0 < float(v) < 20):
            issues.append(f"{yield_field}={v} looks implausible (expected 0-20%)")
            missing.append(yield_field)  # treat as missing — will attempt repair

    if not result.get("market_news"):
        issues.append("market_news is empty")
    if not result.get("earnings"):
        issues.append("earnings is empty")

    return {
        "ok": not missing and not issues,
        "missing_fixed_income": sorted(set(missing)),
        "issues": issues,
    }

async def _agent_repair(api_key: str, missing_fields: list) -> dict:
    """REPAIR step — one small, targeted follow-up call asking ONLY for the
    fields that came back missing/implausible. Cheap (short prompt, few
    searches) compared to re-running the full fetch.
    """
    if not missing_fields:
        return {}
    field_hints = {
        "gsec_10y": 'search "India 10 year government bond yield today"',
        "gsec_5y": 'search "India 5 year government bond yield today"',
        "gsec_1y": 'search "India 1 year treasury bill yield rate"',
        "repo_rate": 'search "RBI repo rate current"',
        "rbi_stance": 'search "RBI monetary policy stance latest"',
        "cpi_inflation": 'search "India CPI inflation rate latest month"',
        "aaa_spread_10y": 'search "India AAA corporate bond yield" and subtract gsec_10y, result in bps',
        "mibor_overnight": 'search "India MIBOR overnight rate" or "overnight call money rate"',
        "yield_curve_slope": "calculate (gsec_10y - gsec_1y) * 100 in bps from already-known yields",
    }
    lines = [f"- {f}: {field_hints.get(f, 'search for ' + f)}" for f in missing_fields]
    schema = {f: (0.0 if f != "rbi_stance" else "") for f in missing_fields}
    prompt = (
        f"Find ONLY these specific India fixed-income data points, nothing else:\n"
        + "\n".join(lines) +
        f"\n\nReturn ONLY this JSON with the values you found (use null only if truly unavailable "
        f"after searching): {json.dumps(schema)}"
    )
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 500,
                    "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 4}],
                    "messages": [{"role": "user", "content": prompt}]
                })
            if resp.status_code != 200:
                log.warning(f"Agent repair HTTP {resp.status_code}: {resp.text[:150]}")
                return {}
            data = resp.json()
            blocks = data.get("content", [])
            if data.get("stop_reason") == "tool_use":
                tool_results = [
                    {"type": "tool_result", "tool_use_id": b.get("tool_use_id"), "content": b.get("content", [])}
                    for b in blocks if b.get("type") == "web_search_tool_result"
                ]
                resp2 = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 400,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 4}],
                        "messages": [
                            {"role": "user", "content": prompt},
                            {"role": "assistant", "content": blocks},
                            {"role": "user", "content": tool_results if tool_results else
                             [{"type": "text", "text": "Provide the JSON now based on search results."}]}
                        ]
                    })
                blocks = resp2.json().get("content", []) if resp2.status_code == 200 else blocks
            text = " ".join(b.get("text", "") for b in blocks if b.get("type") == "text").strip()
            text = re.sub(r"```[^\n]*\n?|```", "", text).strip()
            start, end = text.find("{"), text.rfind("}") + 1
            if start >= 0 and end > start:
                return json.loads(text[start:end])
    except Exception as e:
        log.warning(f"Agent repair failed: {e}")
    return {}

async def _fetch_news_and_earnings(api_key: str, stock_list: str) -> dict:
    """Agent entry point — PLAN, ACT, VERIFY, REPAIR, REPORT."""
    import time
    if _news_cache["data"] and (time.time() - _news_cache["ts"]) < 21600:  # 6 hour cache
        return _news_cache["data"]

    cycle = {"stage": "act", "repaired_fields": [], "fallback_fields": [], "errors": []}
    today = date.today().strftime("%d %B %Y")
    prompt = (
        f"Today is {today}. Search the web for the MOST RECENT Indian market news, earnings, "
        f"and fixed income data you can find — it does not need to be from today specifically, "
        f"the most recent available data within the last 2-4 weeks is perfectly fine.\n\n"
        f"Run SEPARATE, SPECIFIC searches for each of these fixed income data points — do not "
        f"combine them into one vague search:\n"
        f"1. \"India 10 year government bond yield today\" — for gsec_10y\n"
        f"2. \"India 5 year government bond yield today\" — for gsec_5y\n"
        f"3. \"India 1 year treasury bill yield rate\" — for gsec_1y\n"
        f"4. \"RBI repo rate current\" — for repo_rate and rbi_stance\n"
        f"5. \"India CPI inflation rate latest month\" — for cpi_inflation\n"
        f"6. \"India AAA corporate bond spread over government bond\" or \"AAA bond yield India\" — for aaa_spread_10y "
        f"(if you find AAA corporate bond yield but not the spread, calculate spread = AAA yield - gsec_10y, in basis points)\n"
        f"7. \"India MIBOR overnight rate today\" or \"India overnight call money rate\" — for mibor_overnight\n\n"
        f"After finding gsec_10y and gsec_1y, calculate yield_curve_slope = (gsec_10y - gsec_1y) * 100, in basis points.\n\n"
        f"IMPORTANT: You must always return the JSON below, using your best available recent data. "
        f"Never apologize, never explain limitations, never refuse — just fill in the JSON with the "
        f"most recent real figures you found, even if approximate or slightly dated. Only use null if "
        f"you genuinely found nothing after searching — but try all 7 searches above before giving up "
        f"on any field.\n\n"
        f"For beat_miss: use \"Miss\" whenever revenue_growth_pct OR profit_growth_pct is "
        f"NEGATIVE (an actual year-over-year decline), even if the company technically beat a "
        f"lowered analyst estimate. Do not label a real earnings decline as \"Beat\" or "
        f"\"In-line\" just because it cleared a reduced forecast — the advisor needs to see the "
        f"real decline. Use \"Beat\" only for genuine positive YoY growth that exceeded "
        f"expectations, \"In-line\" for flat/as-expected results, \"Miss\" for any negative "
        f"YoY decline or a result below expectations.\n\n"
        "Return ONLY the JSON object, no other text:\n"
        '{"earnings":[{"company":"","result_date":"","revenue_growth_pct":0,"profit_growth_pct":0,"beat_miss":"Beat|In-line|Miss"}],'
        '"market_news":[{"headline":"","category":"Market","sentiment":"Positive|Neutral|Negative"}],'
        '"portfolio_news":[{"stock":"","headline":"","sentiment":"Positive|Neutral|Negative"}],'
        '"fixed_income":{"gsec_10y":0.0,"gsec_5y":0.0,"gsec_1y":0.0,"repo_rate":0.0,"rbi_stance":"","cpi_inflation":0.0,"aaa_spread_10y":0,"mibor_overnight":0.0,"yield_curve_slope":0,"debt_market_view":""}}\n\n'
        f"Include: 3 earnings, 4 market news, 3 stock news items for {stock_list[:50]}, and all 9 fixed income fields. "
        f"Remember: output ONLY the JSON, with your best recent data filled in — no apology text, no caveats."
    )

    result = {"earnings": [], "market_news": [], "portfolio_news": [], "fixed_income": {}}
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            # ── ACT: step 1 — call with web_search, retry once on rate limit ──
            resp1 = None
            for attempt in range(2):
                resp1 = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 2000,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 7,
                                   "user_location": {"type": "approximate", "city": "Chennai",
                                                     "region": "Tamil Nadu", "country": "IN",
                                                     "timezone": "Asia/Kolkata"}}],
                        "messages": [{"role": "user", "content": prompt}]
                    })
                if resp1.status_code == 429 and attempt == 0:
                    log.warning("Agent: rate limited, retrying in 15s...")
                    cycle["errors"].append("rate_limited_retry")
                    await asyncio.sleep(15)
                    continue
                break

            if resp1.status_code != 200:
                cycle["errors"].append(f"step1_http_{resp1.status_code}")
                log.warning(f"Agent step1 HTTP {resp1.status_code}: {resp1.text[:200]}")
                _agent_log({**cycle, "stage": "failed", "success": False})
                return result

            resp1_data = resp1.json()
            assistant_content = resp1_data.get("content", [])

            if resp1_data.get("stop_reason") == "tool_use":
                tool_results = [
                    {"type": "tool_result", "tool_use_id": b.get("tool_use_id"), "content": b.get("content", [])}
                    for b in assistant_content if b.get("type") == "web_search_tool_result"
                ]
                resp2 = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 2500,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 7}],
                        "messages": [
                            {"role": "user", "content": prompt},
                            {"role": "assistant", "content": assistant_content},
                            {"role": "user", "content": tool_results if tool_results else
                             [{"type": "text", "text": "Based on the search results above, provide the JSON."}]}
                        ]
                    })
                final_blocks = resp2.json().get("content", []) if resp2.status_code == 200 else assistant_content
            else:
                final_blocks = assistant_content

            text = " ".join(b.get("text", "") for b in final_blocks if b.get("type") == "text").strip()
            text = re.sub(r"```[^\n]*\n?|```", "", text).strip()
            start, end = text.find("{"), text.rfind("}") + 1
            if start >= 0 and end > start:
                result = json.loads(text[start:end])
            else:
                cycle["errors"].append("no_json_in_response")
                log.warning(f"Agent: no JSON found. Text: {text[:300]}")

            # ── VERIFY ──
            cycle["stage"] = "verify"
            verdict = _agent_verify(result)
            cycle["issues"] = verdict["issues"]

            # ── REPAIR — one targeted follow-up for missing/implausible fields ──
            if verdict["missing_fixed_income"]:
                cycle["stage"] = "repair"
                log.info(f"Agent: repairing {verdict['missing_fixed_income']}")
                repaired = await _agent_repair(api_key, verdict["missing_fixed_income"])
                if repaired:
                    fi = result.setdefault("fixed_income", {})
                    for k, v in repaired.items():
                        if v is not None and k in verdict["missing_fixed_income"]:
                            fi[k] = v
                            cycle["repaired_fields"].append(k)

            # ── Final fallback — only for fields STILL missing after repair ──
            verdict2 = _agent_verify(result)
            if verdict2["missing_fixed_income"]:
                fi = result.setdefault("fixed_income", {})
                for k in verdict2["missing_fixed_income"]:
                    if k in FIXED_INCOME_FALLBACKS and fi.get(k) is None:
                        fi[k] = FIXED_INCOME_FALLBACKS[k]
                        cycle["fallback_fields"].append(k)
                fi["_estimated_fields"] = cycle["fallback_fields"]  # client can show this if needed

            cycle["success"] = True
            cycle["news_count"] = len(result.get("market_news", []))
            cycle["earnings_count"] = len(result.get("earnings", []))
            _agent_log(cycle)

            _news_cache["data"] = result
            _news_cache["ts"] = time.time()
            log.info(f"Agent OK: {cycle['news_count']} news, {cycle['earnings_count']} earnings, "
                     f"repaired={cycle['repaired_fields']}, fallback={cycle['fallback_fields']}")
            return result

    except Exception as e:
        cycle["errors"].append(str(e)[:150])
        cycle["success"] = False
        _agent_log(cycle)
        log.warning(f"Agent fetch failed: {e}")
    return result

@app.get("/agent-health")
async def agent_health():
    """Visible health log — last N fetch cycles with what happened at each stage."""
    return {
        "cycles_logged": len(_agent_health),
        "history": _agent_health,
    }

@app.get("/debug-news-raw")
async def debug_news_raw():
    """Debug: expose raw HTTP status/body at each step of the news fetch chain."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"error": "ANTHROPIC_API_KEY not set"}
    import json as _j
    from datetime import date
    today = date.today().strftime("%d %B %Y")
    prompt = (
        f"Today is {today}. Search the web for the MOST RECENT Indian market news, earnings, "
        f"and fixed income data you can find — it does not need to be from today specifically, "
        f"the most recent available data within the last 1-2 weeks is perfectly fine.\n\n"
        f"IMPORTANT: You must always return the JSON below, using your best available recent data. "
        f"Never apologize, never explain limitations, never refuse — just fill in the JSON with the "
        f"most recent real figures you found, even if approximate or slightly dated. If something is "
        f"truly unavailable use null, but still output the full JSON structure.\n\n"
        "Return ONLY the JSON object, no other text:\n"
        '{"earnings":[{"company":"","result_date":"","revenue_growth_pct":0,"profit_growth_pct":0,"beat_miss":"Beat"}],'
        '"market_news":[{"headline":"","category":"Market","sentiment":"Positive"}],'
        '"portfolio_news":[{"stock":"","headline":"","sentiment":"Positive"}],'
        '"fixed_income":{"gsec_10y":0.0,"gsec_1y":0.0,"repo_rate":0.0,"rbi_stance":"","cpi_inflation":0.0,"aaa_spread_10y":0,"debt_market_view":""}}\n\n'
        "Include: 3 earnings, 4 market news, 3 stock news items for HDFC Bank Reliance TCS, and fixed income data. "
        "Remember: output ONLY the JSON, with your best recent data filled in — no apology text, no caveats."
    )
    debug = {}
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            resp1 = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 4000,
                    "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 8}],
                    "messages": [{"role": "user", "content": prompt}]
                })
            debug["step1_status"] = resp1.status_code
            if resp1.status_code != 200:
                debug["step1_error"] = resp1.text[:500]
                return debug
            r1 = resp1.json()
            debug["step1_stop_reason"] = r1.get("stop_reason")
            debug["step1_block_types"] = [b.get("type") for b in r1.get("content", [])]
            assistant_content = r1.get("content", [])

            if r1.get("stop_reason") == "tool_use":
                tool_results = []
                for block in assistant_content:
                    if block.get("type") == "web_search_tool_result":
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.get("tool_use_id"),
                            "content": block.get("content", [])
                        })
                debug["tool_results_count"] = len(tool_results)
                resp2 = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 2500,
                        "tools": [{"type": "web_search_20250305", "name": "web_search", "max_uses": 8}],
                        "messages": [
                            {"role": "user", "content": prompt},
                            {"role": "assistant", "content": assistant_content},
                            {"role": "user", "content": tool_results if tool_results else
                             [{"type": "text", "text": "Based on the search results above, provide the JSON."}]}
                        ]
                    })
                debug["step2_status"] = resp2.status_code
                if resp2.status_code != 200:
                    debug["step2_error"] = resp2.text[:500]
                    return debug
                r2 = resp2.json()
                debug["step2_stop_reason"] = r2.get("stop_reason")
                debug["step2_block_types"] = [b.get("type") for b in r2.get("content", [])]
                final_blocks = r2.get("content", [])
            else:
                final_blocks = assistant_content

            text = " ".join(b.get("text", "") for b in final_blocks if b.get("type") == "text").strip()
            debug["raw_text_len"] = len(text)
            debug["raw_text_preview"] = text[:600]
            cleaned = re.sub(r"```[^\n]*\n?|```", "", text).strip()
            start = cleaned.find("{")
            end = cleaned.rfind("}") + 1
            debug["json_found"] = start >= 0 and end > start
            if debug["json_found"]:
                try:
                    parsed = _j.loads(cleaned[start:end])
                    debug["parsed_keys"] = list(parsed.keys())
                    debug["news_count"] = len(parsed.get("market_news", []))
                except Exception as e:
                    debug["parse_error"] = str(e)
    except Exception as e:
        debug["exception"] = str(e)
    return debug

@app.get("/debug-news")
async def debug_news():
    """Force a fresh news fetch — clears cache and returns result."""
    global _news_cache
    _news_cache = {"data": None, "ts": 0.0}
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"error": "ANTHROPIC_API_KEY not set"}
    result = await _fetch_news_and_earnings(api_key, "HDFC Bank, Reliance, TCS, Infosys, ICICI Bank")
    return {
        "status": "ok",
        "news_count": len(result.get("market_news", [])),
        "earnings_count": len(result.get("earnings", [])),
        "portfolio_news_count": len(result.get("portfolio_news", [])),
        "fixed_income": result.get("fixed_income", {}),
        "sample_news": result.get("market_news", [])[:2],
        "sample_earnings": result.get("earnings", [])[:2],
    }


@app.get("/debug-yahoo")
async def debug_yahoo():
    """Debug: test each Yahoo Finance index symbol individually."""
    global _indices_cache
    _indices_cache = {"data": None, "ts": 0.0}
    test_symbols = [
        ("NIFTY 50", "%5ENSEI"),
        ("NIFTY BANK", "%5ENSEBANK"),
        ("NIFTY MIDCAP 150", "NIFTY_MIDCAP_150.NS"),
        ("NIFTY SMALLCAP 250", "NIFTY_SMLCAP_250.NS"),
        ("INDIA VIX", "%5EINDIAVIX"),
        ("NIFTY IT", "%5ECNXIT"),
    ]
    results = {}
    async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
        for label, symbol in test_symbols:
            try:
                r = await client.get(
                    f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}",
                    params={"range": "5d", "interval": "1d"},
                    headers=YAHOO_HEADERS, timeout=8.0)
                results[label] = {"status": r.status_code, "body": r.text[:300]}
            except Exception as e:
                results[label] = {"error": str(e)}
    return results

@app.get("/debug-indices")
async def debug_indices():
    """Debug: show raw NSE response for the indices endpoint."""
    global _indices_cache
    _indices_cache = {"data": None, "ts": 0.0}  # clear cache to force fresh fetch
    results = {}
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
            session_resp = await client.get("https://www.nseindia.com", headers=NSE_HEADERS, timeout=10.0)
            results["session_status"] = session_resp.status_code
            results["session_cookies"] = list(client.cookies.keys())

            r = await client.get(
                "https://www.nseindia.com/api/equity-stockIndices?index=NIFTY%2050",
                headers=NSE_HEADERS, timeout=8.0)
            results["index_status"] = r.status_code
            results["index_raw"] = r.text[:500]
    except Exception as e:
        results["exception"] = str(e)
    return results

@app.get("/debug-fii")
async def debug_fii():
    """Debug endpoint — shows raw NSE FII/DII response for field inspection."""
    results = {}
    async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
        await _get_nse_session(client)
        for url in [
            "https://www.nseindia.com/api/fiidiiTradeReact",
            "https://www.nseindia.com/api/fii-dii-data",
        ]:
            try:
                r = await client.get(url, headers=NSE_HEADERS, timeout=8.0)
                results[url] = {
                    "status": r.status_code,
                    "raw": r.text[:500] if r.status_code == 200 else r.text[:200]
                }
            except Exception as e:
                results[url] = {"error": str(e)}
    return results

@app.get("/market-test")
async def market_test():
    import time
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key: return {"ok": False, "error": "ANTHROPIC_API_KEY not set"}
    t0 = time.time()
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={"model": "claude-haiku-4-5-20251001", "max_tokens": 10,
                      "messages": [{"role": "user", "content": "Say OK"}]})
        ms = int((time.time()-t0)*1000)
        if resp.status_code == 200:
            return {"ok": True, "ms": ms, "model": "claude-haiku-4-5-20251001"}
        return {"ok": False, "status": resp.status_code, "error": resp.text[:100]}
    except Exception as e:
        return {"ok": False, "error": str(e)[:100]}

@app.get("/market-data")
async def market_data(stocks: str = ""):
    import asyncio, time
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    stock_list = stocks or "HDFC Bank, Reliance, Infosys, ICICI Bank, Axis Bank, TCS, Kotak Mahindra Bank"

    # Fetch indices and FII/DII in parallel (fast, direct from NSE)
    indices, fii_dii = await asyncio.gather(
        _fetch_indices(),
        _fetch_fii_dii(),
        return_exceptions=True
    )
    if isinstance(indices, Exception):
        log.warning(f"Indices fetch error: {indices}")
        indices = []
    if isinstance(fii_dii, Exception):
        log.warning(f"FII/DII fetch error: {fii_dii}")
        fii_dii = {}

    # Fetch news async — use cached if fresh, else fetch in background
    news = {}
    if api_key:
        age = time.time() - _news_cache["ts"]
        if _news_cache["data"] and age < 21600:
            news = _news_cache["data"]
        elif _market_cache["status"] != "fetching":
            _market_cache["status"] = "fetching"
            asyncio.create_task(_fetch_news_bg(api_key, stock_list))
            news = _news_cache.get("data") or {}
    else:
        news = {}

    return {
        "indices":        indices,
        "fii_dii":        fii_dii,
        "earnings":       news.get("earnings", []),
        "market_news":    news.get("market_news", []),
        "portfolio_news": news.get("portfolio_news", []),
        "fixed_income":   news.get("fixed_income", {}),
        "news_status":    "ready" if news else "loading",
    }

async def _fetch_news_bg(api_key: str, stock_list: str):
    """Background task to fetch news without blocking the response."""
    try:
        await _fetch_news_and_earnings(api_key, stock_list)
    except Exception as e:
        log.warning(f"Background news fetch failed: {e}")
    finally:
        _market_cache["status"] = "idle"


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0",
                port=int(os.environ.get("PORT", 8000)), reload=False)
